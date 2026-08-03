"""Authenticated analytical report views."""
from io import BytesIO
from math import floor
from collections import defaultdict

from django.core.exceptions import ImproperlyConfigured
from django.db import DatabaseError
from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..filters import CompletedTransectFilterSet
from ..forms_reports import BoneCensusExportForm, MNIReportForm
from ..models import CompletedTransect
from ..reports.mni_service import build_report
from .mixins import BonesAuthMixin


def _value(cell, key):
    if cell.get("excluded"):
        return "Excluded"
    return cell.get(key)


def export_workbook(result, filters, methodology, transect_names=None):
    transect_names = transect_names or {}
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    header = ["Taxon", "All occurrence n", "All occurrence %", "All MNI n", "All MNI %"]
    for habitat in result.habitats:
        header.extend([f"{habitat} occurrence n", f"{habitat} occurrence %", f"{habitat} MNI n", f"{habitat} MNI %"])
    summary.append(header)
    for cell in summary[1]:
        cell.font = Font(bold=True)
    for row in result.rows:
        values = [row["taxon"], _value(row["occurrence_all"], "count"), _value(row["occurrence_all"], "percent"), _value(row["mni_all"], "count"), _value(row["mni_all"], "percent")]
        for habitat in row["habitats"]:
            values.extend([_value(habitat["occurrence"], "count"), _value(habitat["occurrence"], "percent"), _value(habitat["mni"], "count"), _value(habitat["mni"], "percent")])
        summary.append(values)
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    transects = workbook.create_sheet("Transect MNI")
    transects.append(["Transect UID", "Transect", "Habitat", "Taxon", "MNI"])
    for row in result.transect_rows:
        transects.append([row["transect_id"], transect_names.get(row["transect_id"], ""), row["habitat"], row["taxon"], row["mni"]])
    groups = workbook.create_sheet("Group MNI")
    groups.append(["Transect UID", "Transect", "Taxon", "Sex", "Age", "Weathering", "Group MNI"])
    for row in result.group_rows:
        groups.append([row.transect_id, transect_names.get(row.transect_id, ""), row.taxon, row.sex, row.age, row.weathering, row.mni])
    warnings = workbook.create_sheet("Warnings")
    warnings.append(["Severity", "Category", "Workflow", "Transect UID", "Transect", "Occurrence ID", "Instance", "Contributing instances", "Taxon", "Habitat", "Element / side", "Occurrence n", "MNI n", "Raw value", "Treatment"])
    for row in result.warnings:
        warnings.append([row.get("severity", "Warning"), row.get("category"), row.get("workflow", ""), row.get("transect_id"), transect_names.get(row.get("transect_id"), ""), row.get("occurrence_id"), row.get("instance_number"), ", ".join(map(str, row.get("instance_numbers", []))), row.get("taxon", ""), row.get("habitat", ""), row.get("element_side", ""), row.get("occurrence_n"), row.get("mni_n"), row.get("raw_value"), row.get("treatment")])
    notes = workbook.create_sheet("Methodology")
    notes.append(["MNI report methodology"])
    notes["A1"].font = Font(bold=True)
    for line in methodology:
        notes.append([line])
    notes.append(["Applied filters", filters])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


ANALYSIS_HEADERS = [
    "Date", "Year", "Month", "Property", "Sector", "Block", "Species",
    "Total", "Category", "Remarks", "ObjectId", "GlobalID",
    "CreationDate", "Creator", "EditDate", "Editor",
]
ANALYSIS_KEY = [
    ("Date", "Transect date"),
    ("Year", "Estimated year of death: transect year minus the maximum age for the weathering class"),
    ("Month", "NA"),
    ("Property", "Set to Ol Pejeta"),
    ("Sector", "Eastern = reserve, Western = ranch"),
    ("Block", "The transect name"),
    ("Species", "Canonical taxon name"),
    ("Total", "Estimated MNI"),
    ("Category", "Sex-Age-Bone weathering class; for example Male-adult-3-4"),
    ("Remarks", "OPC vegetation"),
    ("ObjectId", "NA"), ("GlobalID", "NA"),
    ("CreationDate", "Dataset creation date"), ("Creator", "Dataset creator"),
    ("EditDate", "NA"), ("Editor", "NA"),
]


def analysis_filter_rows(cleaned_data, note_filters=()):
    selected_transects = list(cleaned_data.get("transects") or [])
    rows = [
        ("Transects", ", ".join(str(row) for row in selected_transects) or "All completed transects"),
        ("Transect years", ", ".join(cleaned_data.get("years") or []) or "All"),
        ("Habitats", ", ".join(cleaned_data.get("habitats") or []) or "All"),
        ("Old reserve", cleaned_data.get("reserve") or "Both"),
        ("Excluded taxa", ", ".join(cleaned_data.get("excluded_taxa") or []) or "None"),
        ("Include elements", "Yes" if cleaned_data.get("include_elements") else "No"),
        ("Omit unknown weathering", "Yes" if cleaned_data.get("omit_unknown_weathering") else "No"),
        ("Use normalised weathering", "Yes" if cleaned_data.get("use_normalised_weathering") else "No"),
    ]
    if cleaned_data.get("use_normalised_weathering"):
        rows.append((
            "2011 treatment",
            "Original weathering class and original maximum age",
        ))
    for index, row in enumerate(note_filters, start=1):
        rows.append((
            f"Transect note {index}",
            f'{row.get("phase") or "Any phase"} / {row.get("question") or "Any question"}: '
            f'{", ".join(row.get("responses", [])) or "any response"}',
        ))
    return rows


def _group_key(item):
    return (item.transect_id, item.taxon, item.sex, item.age, item.weathering)


def build_group_evidence(result):
    contributing = defaultdict(set)
    for observation in result.usable_observations:
        contributing[_group_key(observation)].add(observation.occurrence_id)

    element_instances = defaultdict(lambda: defaultdict(set))
    for observation in result.observations:
        key = _group_key(observation)
        if observation.occurrence_id not in contributing[key]:
            continue
        element = "tooth" if observation.element.casefold() == "teeth" else observation.element
        element_instances[key][element.casefold()].add(
            (observation.occurrence_id, observation.instance_number)
        )

    evidence = {}
    for key, occurrence_ids in contributing.items():
        elements = []
        for element, instances in sorted(element_instances[key].items()):
            elements.append(element if len(instances) == 1 else f"{element}({len(instances)})")
        occurrences = []
        for occurrence_id in sorted(occurrence_ids):
            number = result.occurrence_metadata.get(occurrence_id, {}).get("number")
            occurrences.append(
                f"{occurrence_id} ({number})" if number is not None else str(occurrence_id)
            )
        evidence[key] = {"elements": ", ".join(elements),
                         "occurrences": ", ".join(occurrences)}
    return evidence


def analysis_export_filename(include_elements):
    return "bone-census-data-elements.xlsx" if include_elements else "bone-census-data.xlsx"


def export_analysis_workbook(result, creator, filters=(), created_on=None,
                             omit_unknown_weathering=False,
                             use_normalised_weathering=False,
                             include_elements=False):
    created_on = created_on or timezone.localdate()
    workbook = Workbook()
    key = workbook.active
    key.title = "key"
    key_rows = list(ANALYSIS_KEY)
    if include_elements:
        remarks_index = next(index for index, row in enumerate(key_rows) if row[0] == "Remarks")
        key_rows[remarks_index:remarks_index] = [
            ("Elements", "Unique contributing elements; repeated instances show a count in parentheses"),
            ("Occurrence", "Contributing occurrence primary key followed by occurrence number in parentheses"),
        ]
    for row in key_rows:
        key.append(row)
    if use_normalised_weathering:
        key["B2"] = (
            "Estimated year of death: transect year minus corrected maximum age "
            "for normalised weathering, except 2011 transects use the original class and maximum age"
        )
        key["B9"] = (
            "Sex-Age-Weathering; normalised exports use 0-2 or 3-5, "
            "except 2011 transects retain the original weathering class"
        )
    key.append([])
    key.append(["Applied filters"])
    key[key.max_row][0].font = Font(bold=True)
    for label, value in filters:
        key.append([label, value])
    data = workbook.create_sheet("Data")
    headers = list(ANALYSIS_HEADERS)
    if include_elements:
        remarks_index = headers.index("Remarks")
        headers[remarks_index:remarks_index] = ["Elements", "Occurrence"]
    data.append(headers)
    for cell in data[1]:
        cell.font = Font(bold=True)
    group_evidence = build_group_evidence(result) if include_elements else {}
    for group in result.group_rows:
        if omit_unknown_weathering and group.weathering.casefold() == "unknown":
            continue
        metadata = result.transect_metadata.get(group.transect_id, {})
        transect_date = metadata.get("date")
        if transect_date and timezone.is_aware(transect_date):
            transect_date = timezone.localtime(transect_date)
        excel_date = transect_date.date() if hasattr(transect_date, "date") else transect_date
        weathering = group.weathering
        age_max = result.weathering_age_max.get(group.weathering.casefold())
        if use_normalised_weathering and transect_date and transect_date.year != 2011:
            normalised = result.weathering_normalised.get(group.weathering.casefold())
            if normalised:
                weathering = normalised["class"]
                age_max = normalised["age_max"]
        estimated_year = (
            floor(transect_date.year - float(age_max))
            if transect_date and age_max is not None else None
        )
        reserve = str(metadata.get("reserve") or "").strip().casefold()
        sector = "Eastern" if reserve == "yes" else "Western" if reserve == "no" else ""
        values = [
            excel_date, estimated_year, None, "Ol Pejeta", sector,
            metadata.get("block", ""), group.taxon, group.mni,
            f"{group.sex}-{group.age}-{weathering}",
        ]
        if include_elements:
            evidence = group_evidence.get(_group_key(group), {})
            values.extend([evidence.get("elements", ""), evidence.get("occurrences", "")])
        values.extend([
            metadata.get("habitat", ""), None, None, created_on, creator,
            None, None,
        ])
        data.append(values)
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions
    widths = {"A": 20, "B": 12, "D": 14, "E": 12, "F": 18, "G": 28,
              "H": 10, "I": 28, "J": 20, "M": 16, "N": 24}
    for column, width in widths.items():
        data.column_dimensions[column].width = width
    key.column_dimensions["A"].width = 18
    key.column_dimensions["B"].width = 90
    for cell in data["A"][1:]:
        cell.number_format = "dd/mm/yyyy"
    creation_column = get_column_letter(headers.index("CreationDate") + 1)
    for cell in data[creation_column][1:]:
        cell.number_format = "yyyy-mm-dd"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class MNISummaryView(BonesAuthMixin, TemplateView):
    template_name = "bones/reports/mni_summary.html"

    methodology = [
        "MNI is calculated independently per transect and then summed.",
        "Only Completed transects and occurrences are included.",
        "All 2008 transects are excluded; in 2024 only shrubs closed habitat is retained.",
        "Bone and Dentition workflows are eligible; Dentition is treated as teeth.",
        "Post taxon, sex and age are preferred, falling back to pre values.",
        "The highest original weathering class recorded in an occurrence is used without analytical regrouping.",
        "Weathering is validated for Bone workflows only; Dentition workflows are not expected to record Weathering class.",
        "Any calculated MNI greater than its occurrence count is reported as a Critical data-quality warning.",
        "Incomplete instances with the same occurrence, taxon, demographic, weathering, element, and recorded side count once; an uncertain-side incomplete fragment does not add another individual when a matching known side is present.",
        "For complete instances, uncertain sides are balanced to give the minimum defensible number of paired elements.",
    ]

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        if context.get("result") and request.GET.get("export") == "xlsx":
            transect_names = {
                transect.pk: getattr(transect.transect_template, "name", "")
                for transect in context["report_transects"]
            }
            payload = export_workbook(context["result"], context["filter_summary"], context["methodology"], transect_names)
            response = HttpResponse(payload, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="mni-summary.xlsx"'
            return response
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        has_filters = any(key != "export" for key in self.request.GET)
        if has_filters:
            data = self.request.GET
        else:
            defaults = MNIReportForm()
            data = self.request.GET.copy()
            data.setlist("excluded_taxa", defaults.initial.get("excluded_taxa", []))
        form = MNIReportForm(data=data)
        note_filter = CompletedTransectFilterSet(data=data, queryset=CompletedTransect.objects.none())
        context.update({"page_title": "MNI summary", "form": form,
                        "note_filter": note_filter, "methodology": self.methodology,
                        "result": None, "report_error": "",
                        "export_query": self.request.GET.urlencode() + ("&" if self.request.GET else "") + "export=xlsx"})
        if form.is_valid():
            try:
                result, transects, excluded = build_report(form.cleaned_data, note_filter.note_filters)
                result.warnings.sort(key=lambda warning: (
                    0 if warning.get("severity") == "Critical" else 1,
                    warning.get("category", ""),
                    str(warning.get("transect_id") or ""),
                    str(warning.get("occurrence_id") or ""),
                ))
                transect_names = {
                    transect.pk: getattr(transect.transect_template, "name", "")
                    for transect in transects
                }
                for warning in result.warnings:
                    warning["transect"] = transect_names.get(warning.get("transect_id"), "")
                    if warning.get("occurrence_id") and warning.get("instance_number"):
                        warning["url"] = reverse("bones:occurrences:instance_detail", kwargs={"occurrence_pk": warning["occurrence_id"], "instance_number": warning["instance_number"]})
                    elif warning.get("occurrence_id"):
                        warning["url"] = reverse("bones:occurrences:detail", kwargs={"pk": warning["occurrence_id"]})
                    elif warning.get("transect_id"):
                        warning["url"] = reverse("bones:transects:detail", kwargs={"pk": warning["transect_id"]})
                selected = form.cleaned_data
                parts = [f"{len(transects)} eligible transects"]
                parts.append("years: " + (", ".join(selected.get("years", [])) or "all eligible"))
                parts.append("habitats: " + (", ".join(selected.get("habitats", [])) or "all"))
                parts.append("old reserve: " + (selected.get("reserve") or "both"))
                if note_filter.note_filters:
                    note_text = [f'{row.get("phase")}/{row.get("question")}: {", ".join(row.get("responses", [])) or "any response"}' for row in note_filter.note_filters]
                    parts.append("transect notes: " + "; ".join(note_text))
                parts.append("excluded taxa: " + (", ".join(excluded) or "none"))
                summary = "; ".join(parts)
                context.update({"result": result, "report_transects": transects, "filter_summary": summary})
            except (DatabaseError, ImproperlyConfigured):
                context["report_error"] = "The report database is temporarily unavailable. Please try again later."
        return context


class MNIAnalysisExportView(BonesAuthMixin, TemplateView):
    template_name = "bones/reports/mni_analysis_export.html"

    def _form_data(self):
        data = self.request.GET.copy()
        if not self.request.GET:
            data["omit_unknown_weathering"] = "on"
            data["use_normalised_weathering"] = "on"
        return data

    def get(self, request, *args, **kwargs):
        data = self._form_data()
        form = BoneCensusExportForm(data=data, apply_population_rules=False)
        note_filter = CompletedTransectFilterSet(data=data, queryset=CompletedTransect.objects.none())
        if request.GET.get("export") == "xlsx" and form.is_valid():
            try:
                result, _, _ = build_report(
                    form.cleaned_data, note_filter.note_filters,
                    apply_population_rules=False,
                )
                creator = request.user.get_full_name().strip() or request.user.get_username()
                filter_rows = analysis_filter_rows(form.cleaned_data, note_filter.note_filters)
                payload = export_analysis_workbook(
                    result, creator, filter_rows,
                    omit_unknown_weathering=form.cleaned_data["omit_unknown_weathering"],
                    use_normalised_weathering=form.cleaned_data["use_normalised_weathering"],
                    include_elements=form.cleaned_data["include_elements"],
                )
                response = HttpResponse(payload, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                filename = analysis_export_filename(form.cleaned_data["include_elements"])
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            except (DatabaseError, ImproperlyConfigured):
                report_error = "The report database is temporarily unavailable. Please try again later."
        else:
            report_error = ""
        filter_rows = analysis_filter_rows(form.cleaned_data, note_filter.note_filters) if form.is_valid() else []
        return self.render_to_response({
            "page_title": "Bone Census Data", "form": form,
            "note_filter": note_filter, "report_error": report_error,
            "filter_rows": filter_rows,
        })
