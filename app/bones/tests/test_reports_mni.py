from io import BytesIO
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace

from django.test import SimpleTestCase
from django.utils import timezone
from django.urls import reverse
from openpyxl import load_workbook

from ..forms_reports import BoneCensusExportForm
from ..reports.mni import (ElementRule, GroupMNI, Observation, OccurrenceRecord,
                           ReportResult, calculate_mni, element_mni)
from ..reports.mni_service import (_weathering_for_occurrence, normalize_side,
                                   weather_group, weather_score)
from ..views.reports import (analysis_export_filename, analysis_filter_rows,
                             export_analysis_workbook, export_workbook)
from ..mni_seed import (DEFAULT_EXCLUDED_TAXA, ELEMENT_RULES, TAXON_RULES,
                        WEATHERING_CORRECTED_RANGES, WEATHERING_RULES)


class ElementMNITests(SimpleTestCase):
    def test_paired_element_uses_largest_known_side(self):
        rule = ElementRule("femur", 1, True)
        self.assertEqual(element_mni(1, 2, 0, 0, rule), 2)

    def test_three_unknown_paired_elements_mean_two_individuals(self):
        rule = ElementRule("femur", 1, True)
        self.assertEqual(element_mni(0, 0, 3, 0, rule), 2)

    def test_unknown_balances_with_known_sides(self):
        rule = ElementRule("femur", 1, True)
        self.assertEqual(element_mni(1, 2, 3, 0, rule), 3)

    def test_unpaired_element_uses_full_count(self):
        rule = ElementRule("atlas", 1, False)
        self.assertEqual(element_mni(0, 0, 0, 2, rule), 2)

    def test_dental_divisor(self):
        rule = ElementRule("teeth", 32, False)
        self.assertEqual(element_mni(0, 0, 33, 0, rule), 2)


class MNIReportCalculationTests(SimpleTestCase):
    def observation(self, transect, occurrence, instance, **kwargs):
        values = {"taxon": "Aepyceros melampus", "sex": "Unknown", "age": "adult",
                  "weathering": "1-2", "element": "femur", "side": "right",
                  "complete": True, "habitat": "grass closed"}
        values.update(kwargs)
        return Observation(transect, occurrence, instance, **values)

    def test_incomplete_duplicate_fragments_count_once(self):
        observations = [self.observation(1, 10, n, complete=False) for n in (1, 2, 3)]
        result = calculate_mni(observations, [OccurrenceRecord(10, 1, "Aepyceros melampus", "grass closed")],
                               {"femur": ElementRule("femur", 1, True)})
        self.assertEqual(result.transect_rows[0]["mni"], 1)

    def test_incomplete_fragments_are_deduplicated_within_recorded_side(self):
        observations = [
            self.observation(1, 10, 22, complete=False, element="innominate", side="left"),
            self.observation(1, 10, 23, complete=False, element="innominate", side="right"),
            self.observation(1, 10, 24, complete=False, element="innominate", side="unknown"),
            self.observation(1, 10, 25, complete=False, element="innominate", side="unknown"),
        ]
        result = calculate_mni(
            observations,
            [OccurrenceRecord(10, 1, "Equus burchellii", "grass closed")],
            {"innominate": ElementRule("innominate", 1, True)},
        )

        self.assertEqual(result.transect_rows[0]["mni"], 1)
        self.assertFalse(any(
            warning["category"] == "occurrence_mni_exceeds_one"
            for warning in result.warnings
        ))

    def test_complete_unknown_side_is_balanced_as_independent_evidence(self):
        observations = [
            self.observation(1, 10, 22, element="innominate", side="left"),
            self.observation(1, 10, 23, element="innominate", side="right"),
            self.observation(1, 10, 24, element="innominate", side="unknown"),
        ]
        result = calculate_mni(
            observations,
            [OccurrenceRecord(10, 1, "Equus burchellii", "grass closed")],
            {"innominate": ElementRule("innominate", 1, True)},
        )

        self.assertEqual(result.transect_rows[0]["mni"], 2)
        occurrence_warning = next(
            warning for warning in result.warnings
            if warning["category"] == "occurrence_mni_exceeds_one"
        )
        self.assertEqual(occurrence_warning["instance_numbers"], [22, 23, 24])

    def test_complete_records_count_independently(self):
        observations = [self.observation(1, 10, n) for n in (1, 2)]
        result = calculate_mni(observations, [OccurrenceRecord(10, 1, "Aepyceros melampus", "grass closed")],
                               {"femur": ElementRule("femur", 1, True)})
        self.assertEqual(result.transect_rows[0]["mni"], 2)

    def test_two_complete_left_elements_raise_critical_mni_warnings(self):
        observations = [
            self.observation(1, 10, instance, element="scapula", side="left")
            for instance in (1, 2)
        ]
        result = calculate_mni(
            observations,
            [OccurrenceRecord(10, 1, "Aepyceros melampus", "grass closed")],
            {"scapula": ElementRule("scapula", 1, True)},
        )

        categories = [warning["category"] for warning in result.warnings]
        self.assertIn("occurrence_mni_exceeds_one", categories)
        self.assertIn("transect_mni_exceeds_occurrences", categories)
        self.assertIn("summary_mni_exceeds_occurrences", categories)
        occurrence_warning = next(
            warning for warning in result.warnings
            if warning["category"] == "occurrence_mni_exceeds_one"
        )
        self.assertEqual(occurrence_warning["severity"], "Critical")
        self.assertEqual(occurrence_warning["instance_numbers"], [1, 2])
        self.assertIn("scapula", occurrence_warning["element_side"])
        self.assertEqual(occurrence_warning["mni_n"], 2)
        taxon_row = next(row for row in result.rows if row["taxon"] == "Aepyceros melampus")
        self.assertTrue(taxon_row["mni_all"]["warning"])

    def test_left_and_right_pair_remains_one_without_mni_mismatch_warning(self):
        observations = [
            self.observation(1, 10, 1, element="scapula", side="left"),
            self.observation(1, 10, 2, element="scapula", side="right"),
        ]
        result = calculate_mni(
            observations,
            [OccurrenceRecord(10, 1, "Aepyceros melampus", "grass closed")],
            {"scapula": ElementRule("scapula", 1, True)},
        )

        self.assertEqual(result.transect_rows[0]["mni"], 1)
        self.assertFalse(any(
            warning["category"].endswith("mni_exceeds_occurrences")
            or warning["category"] == "occurrence_mni_exceeds_one"
            for warning in result.warnings
        ))

    def test_mni_is_calculated_independently_by_transect(self):
        observations = [self.observation(1, 10, 1), self.observation(2, 20, 1)]
        occurrences = [OccurrenceRecord(10, 1, "Aepyceros melampus", "grass closed"),
                       OccurrenceRecord(20, 2, "Aepyceros melampus", "grass closed")]
        result = calculate_mni(observations, occurrences, {"femur": ElementRule("femur", 1, True)})
        taxon = result.rows[1]
        self.assertEqual(taxon["mni_all"]["count"], 2)
        self.assertEqual(taxon["occurrence_all"]["count"], 2)

    def test_excluded_taxon_remains_in_occurrence_summary(self):
        occurrences = [OccurrenceRecord(10, 1, "Ungulata", "grass closed")]
        result = calculate_mni([], occurrences, {}, ["Ungulata"])
        row = result.rows[1]
        self.assertEqual(row["occurrence_all"]["count"], 1)
        self.assertTrue(row["mni_all"]["excluded"])

    def test_export_has_auditable_sheets(self):
        result = calculate_mni([], [], {})
        data = export_workbook(result, "all eligible transects", ["Rule one"])
        workbook = load_workbook(BytesIO(data))
        self.assertEqual(workbook.sheetnames, ["Summary", "Transect MNI", "Group MNI", "Warnings", "Methodology"])

    def test_export_includes_template_transect_name(self):
        observation = Observation(
            1, 10, 1, "Aepyceros melampus", "Unknown", "adult", "1-2",
            "femur", "right", True, "grass closed",
        )
        occurrence = OccurrenceRecord(10, 1, "Aepyceros melampus", "grass closed")
        warning = {
            "category": "example", "transect_id": 1, "occurrence_id": 10,
            "instance_number": 1, "raw_value": "value", "treatment": "Retained",
        }
        result = calculate_mni(
            [observation], [occurrence],
            {"femur": ElementRule("femur", 1, True)}, warnings=[warning],
        )
        workbook = load_workbook(BytesIO(export_workbook(
            result, "one transect", ["Rule one"], {1: "Template transect A"},
        )))

        for sheet_name in ("Transect MNI", "Group MNI", "Warnings"):
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("Transect", headers)
            self.assertEqual(sheet.cell(2, headers.index("Transect") + 1).value, "Template transect A")
        self.assertIn("Severity", [cell.value for cell in workbook["Warnings"][1]])
        self.assertIn("Workflow", [cell.value for cell in workbook["Warnings"][1]])
        warning_headers = [cell.value for cell in workbook["Warnings"][1]]
        for header in ("Contributing instances", "Taxon", "Element / side", "Occurrence n", "MNI n"):
            self.assertIn(header, warning_headers)

    def test_bone_census_export_matches_contract_and_uses_age_max(self):
        result = ReportResult(
            habitats=[], rows=[], transect_rows=[],
            group_rows=[GroupMNI(1, "Equus burchellii", "Unknown", "adult", "5", 2)],
            transect_metadata={1: {
                "date": timezone.make_aware(datetime(2026, 7, 8, 9, 30)), "block": "48",
                "habitat": "grass closed", "reserve": "Yes",
            }},
            weathering_age_max={"5": Decimal("25")},
        )
        workbook = load_workbook(BytesIO(export_analysis_workbook(
            result, "Test User", [("Habitats", "grass closed")], date(2026, 8, 2),
        )))

        self.assertEqual(workbook.sheetnames, ["key", "Data"])
        data = workbook["Data"]
        self.assertEqual([cell.value for cell in data[1]], [
            "Date", "Year", "Month", "Property", "Sector", "Block",
            "Species", "Total", "Category", "Remarks", "ObjectId",
            "GlobalID", "CreationDate", "Creator", "EditDate", "Editor",
        ])
        self.assertEqual(data["B2"].value, 2001)
        self.assertEqual(data["A2"].value, datetime(2026, 7, 8, 0, 0))
        self.assertEqual(data["A2"].number_format, "dd/mm/yyyy")
        self.assertEqual(data["E2"].value, "Eastern")
        self.assertEqual(data["I2"].value, "Unknown-adult-5")
        self.assertEqual(data["N2"].value, "Test User")
        self.assertIsNone(data["C2"].value)
        self.assertIn("maximum age", workbook["key"]["B2"].value)
        key_values = [tuple(cell.value for cell in row) for row in workbook["key"].iter_rows()]
        self.assertIn(("Habitats", "grass closed"), key_values)

    def test_analysis_filter_rows_describe_unfiltered_export(self):
        rows = dict(analysis_filter_rows({}))
        self.assertEqual(rows["Transects"], "All completed transects")
        self.assertEqual(rows["Excluded taxa"], "None")

    def test_normalised_export_uses_corrected_range_except_for_2011(self):
        result = ReportResult(
            habitats=[], rows=[], transect_rows=[],
            group_rows=[
                GroupMNI(1, "Equus burchellii", "Unknown", "adult", "3", 1),
                GroupMNI(2, "Equus burchellii", "Unknown", "adult", "3", 1),
            ],
            transect_metadata={
                1: {"date": datetime(2018, 8, 1), "block": "A"},
                2: {"date": datetime(2011, 8, 1), "block": "B"},
            },
            weathering_age_max={"3": Decimal("8")},
            weathering_normalised={"3": {"class": "3-5", "age_max": Decimal("25")}},
        )
        workbook = load_workbook(BytesIO(export_analysis_workbook(
            result, "Test User", created_on=date(2026, 8, 2),
            use_normalised_weathering=True,
        )))
        data = workbook["Data"]

        self.assertEqual(data["B2"].value, 1993)
        self.assertEqual(data["I2"].value, "Unknown-adult-3-5")
        self.assertEqual(data["B3"].value, 2003)
        self.assertEqual(data["I3"].value, "Unknown-adult-3")

    def test_unknown_weathering_rows_can_be_omitted(self):
        result = ReportResult(
            habitats=[], rows=[], transect_rows=[],
            group_rows=[GroupMNI(1, "Equus burchellii", "Unknown", "adult", "Unknown", 1)],
            transect_metadata={1: {"date": datetime(2018, 8, 1)}},
        )
        workbook = load_workbook(BytesIO(export_analysis_workbook(
            result, "Test User", omit_unknown_weathering=True,
        )))
        self.assertEqual(workbook["Data"].max_row, 1)

    def test_elements_export_aggregates_instances_and_occurrences(self):
        def observation(occurrence, instance, element, side="left"):
            return Observation(
                1, occurrence, instance, "Equus burchellii", "Unknown",
                "adult", "1", element, side, True, "grass closed",
            )

        rib_left = observation(10, 1, "rib", "left")
        rib_same_instance = observation(10, 1, "rib", "right")
        rib_second_instance = observation(11, 2, "rib", "left")
        tooth = observation(11, 3, "teeth", "not applicable")
        result = ReportResult(
            habitats=[], rows=[], transect_rows=[],
            group_rows=[GroupMNI(1, "Equus burchellii", "Unknown", "adult", "1", 1)],
            observations=[rib_left, rib_same_instance, rib_second_instance, tooth],
            usable_observations=[rib_left, rib_second_instance, tooth],
            transect_metadata={1: {"date": datetime(2018, 8, 1), "block": "70"}},
            occurrence_metadata={10: {"number": 1}, 11: {"number": 3}},
            weathering_age_max={"1": Decimal("2.5")},
        )
        workbook = load_workbook(BytesIO(export_analysis_workbook(
            result, "Test User", include_elements=True,
        )))
        data = workbook["Data"]
        headers = [cell.value for cell in data[1]]

        self.assertEqual(headers[9:12], ["Elements", "Occurrence", "Remarks"])
        self.assertEqual(data["J2"].value, "rib(2), tooth")
        self.assertEqual(data["K2"].value, "10 (1), 11 (3)")
        key_names = [cell.value for cell in workbook["key"]["A"]]
        self.assertIn("Elements", key_names)
        self.assertIn("Occurrence", key_names)

    def test_elements_filename_is_conditional(self):
        self.assertEqual(analysis_export_filename(False), "bone-census-data.xlsx")
        self.assertEqual(analysis_export_filename(True), "bone-census-data-elements.xlsx")

    def test_normalised_filters_include_2011_treatment_note(self):
        rows = dict(analysis_filter_rows({
            "include_elements": True,
            "use_normalised_weathering": True,
        }))
        self.assertEqual(rows["Include elements"], "Yes")
        self.assertEqual(
            rows["2011 treatment"],
            "Original weathering class and original maximum age",
        )


class NormalizationTests(SimpleTestCase):
    def test_selected_transect_objects_are_reduced_to_primary_keys(self):
        from unittest.mock import MagicMock, patch

        queryset = MagicMock()
        queryset.filter.return_value = queryset
        queryset.select_related.return_value = queryset
        queryset.order_by.return_value = queryset
        queryset.distinct.return_value = queryset
        selected = SimpleNamespace(pk=7131446)
        with patch("bones.reports.mni_service.CompletedTransect.objects", queryset):
            from ..reports.mni_service import eligible_transects
            eligible_transects({"transects": [selected]}, apply_population_rules=False)
        queryset.filter.assert_any_call(pk__in=[7131446])

    def test_lr_side_is_uncertain(self):
        self.assertEqual(normalize_side("L;R"), ("unknown", True))

    def test_weathering_uses_highest_number_and_groups(self):
        self.assertEqual(weather_score("3-4"), 4)
        self.assertEqual(weather_group(4), "3-4")
        self.assertEqual(weather_group(None), "Unknown")

    def test_original_weathering_class_is_retained(self):
        workflow = SimpleNamespace(
            pk="b1", instance_number=1,
            template_workflow=SimpleNamespace(name="Bone"),
        )
        rule = SimpleNamespace(source_class="4-5")
        weathering, warnings = _weathering_for_occurrence(
            [workflow], {"b1": {"weathering class": "4-5"}}, 1, 10,
            {"4-5": rule},
        )
        self.assertEqual(weathering, "4-5")
        self.assertEqual(warnings, [])

    def test_dentition_without_weathering_does_not_warn(self):
        workflow = SimpleNamespace(
            pk="d1", instance_number=1,
            template_workflow=SimpleNamespace(name="Dentition"),
        )
        score, warnings = _weathering_for_occurrence(
            [workflow], {"d1": {}}, 1, 10,
        )
        self.assertEqual(score, "Unknown")
        self.assertEqual(warnings, [])

    def test_missing_bone_weathering_warns_for_specific_instance(self):
        missing = SimpleNamespace(
            pk="b1", instance_number=1,
            template_workflow=SimpleNamespace(name="Bone"),
        )
        valid = SimpleNamespace(
            pk="b2", instance_number=2,
            template_workflow=SimpleNamespace(name="Bone"),
        )
        score, warnings = _weathering_for_occurrence(
            [missing, valid], {"b1": {}, "b2": {"weathering class": "3-4"}}, 1, 10,
        )
        self.assertEqual(score, "3-4")
        self.assertEqual(len(warnings), 1)
        self.assertEqual(warnings[0]["instance_number"], 1)
        self.assertEqual(warnings[0]["workflow"], "Bone")
        self.assertEqual(warnings[0]["severity"], "Warning")


class MNIReportSurfaceTests(SimpleTestCase):
    def test_bone_census_options_are_checked_by_default(self):
        self.assertFalse(BoneCensusExportForm.base_fields["include_elements"].initial)
        self.assertTrue(BoneCensusExportForm.base_fields["omit_unknown_weathering"].initial)
        self.assertTrue(BoneCensusExportForm.base_fields["use_normalised_weathering"].initial)

    def test_report_requires_authentication(self):
        url = reverse("bones:reports:mni")
        response = self.client.get(url)
        self.assertRedirects(response, f'{reverse("login")}?next={url}', fetch_redirect_response=False)

    def test_analysis_export_requires_authentication(self):
        url = reverse("bones:reports:mni_analysis")
        response = self.client.get(url)
        self.assertRedirects(response, f'{reverse("login")}?next={url}', fetch_redirect_response=False)

    def test_report_template_has_semantic_and_responsive_structure(self):
        from pathlib import Path
        template = (Path(__file__).resolve().parents[1] / "templates" / "bones" / "reports" / "mni_summary.html").read_text(encoding="utf-8")
        self.assertIn("<section", template)
        self.assertIn("<table", template)
        self.assertIn('scope="colgroup"', template)
        self.assertIn("bones-report-table-wrap", template)
        self.assertIn("fa-file-excel", template)
        self.assertIn("Methodology", template)
        self.assertIn("<th>Transect UID</th><th>Transect</th>", template)
        self.assertIn("warning.transect|default", template)
        self.assertIn("MNI exceeds occurrence count", template)
        self.assertIn("<th>Occurrence n</th><th>MNI n</th>", template)

    def test_bone_census_footer_updates_excluded_taxa_selection(self):
        from pathlib import Path
        template = (Path(__file__).resolve().parents[1] / "templates" / "bones" / "reports" / "mni_analysis_export.html").read_text(encoding="utf-8")
        self.assertIn('id="bone-census-filter-summary"', template)
        self.assertIn("updateAppliedFilters", template)
        self.assertIn("selectedOptions", template)
        self.assertIn("omit_unknown_weathering", template)
        self.assertIn("updateNoteFilters", template)

    def test_versioned_seed_data_contains_workbook_rules(self):
        elements = dict(ELEMENT_RULES)
        self.assertEqual(elements["teeth"], 32)
        self.assertEqual(elements["thoracic vertebra"], 18)
        aliases = {alias.casefold(): canonical for alias, canonical in TAXON_RULES}
        self.assertEqual(aliases["impala"], "Aepyceros melampus")
        self.assertIn("ungulata", DEFAULT_EXCLUDED_TAXA)
        weathering = {source: (canonical, age_max) for source, canonical, _, age_max in WEATHERING_RULES}
        self.assertNotIn("3-5", weathering)
        self.assertEqual(weathering["3-4"], ("3-4", 20))
        self.assertEqual(weathering["5"], ("5", 25))
        self.assertEqual(WEATHERING_CORRECTED_RANGES["0"], (0, 5))
        self.assertEqual(WEATHERING_CORRECTED_RANGES["2"], (0, 5))
        self.assertEqual(WEATHERING_CORRECTED_RANGES["2-3"], (4, 25))
        self.assertEqual(WEATHERING_CORRECTED_RANGES["6"], (4, 25))
