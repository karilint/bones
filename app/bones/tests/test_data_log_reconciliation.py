import json
import tempfile
from collections import defaultdict
from contextlib import ExitStack
from datetime import datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.test import RequestFactory, SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from ..forms_reports import DataReconciliationReportForm
from ..management.commands.reconcile_data_logs import (
    Command, cross_device_occurrence_candidate, has_valid_track_coordinates,
    is_empty_logged_occurrence,
    filter_reconciliation_rows, is_recoverable_track_event,
    logged_parent_transect_cancelled, omit_logged_instance,
    omit_logged_occurrence,
    legacy_track_key,
    merged_occurrence_candidate,
)
from ..navigation import navigation_context
from ..reports.data_log_reconciliation import SHEETS, gps_status, parse_log, summary_rows, workbook_bytes, write_workbook
from ..views.reports import DataReconciliationReportView


class DataLogParserTests(SimpleTestCase):
    def test_nested_json_entities_are_normalized(self):
        payload = {
            "transects": [{
                "TransectUID": 17, "Name": "T-17", "StartTime": "2025-06-01T10:00:00Z",
                "occurrences": [{
                    "OccurrenceID": 91, "OccurrenceNumber": 2, "TransectUID": 17,
                    "workflows": [{"CompletedWorkflowID": "wf-1", "InstanceNumber": 3, "OccurrenceID": 91}],
                }],
                "track": [{"CompletedTransectUID": 17, "Time": "2025-06-01T10:01:00Z", "Lat": 60.1, "Long": 24.9, "isStart": True}],
            }]
        }
        result = parse_log(4, json.dumps(payload))
        self.assertEqual(result.error, "")
        self.assertEqual(result.transects[0]["uid"], 17)
        self.assertEqual(result.occurrences[0]["number"], 2)
        self.assertEqual(result.instances[0]["number"], 3)
        self.assertEqual(len(result.track_points), 1)

    def test_xml_and_malformed_logs_are_reported(self):
        xml = "<Log><Transect><TransectUID>7</TransectUID><Name>A</Name><StartTime>2024-01-01</StartTime></Transect></Log>"
        self.assertEqual(parse_log(1, xml).transects[0]["uid"], 7)
        self.assertTrue(parse_log(2, "not structured data").error)
        self.assertTrue(parse_log(3, "").error)

    def test_bones_line_log_entities_and_cancellations_are_normalized(self):
        payload = "\n".join([
            "STARTTRANSECT template-guid 4354068 2.4,-1.4 17-Feb-2003 16:50:58 1",
            "/STARTTRANSECT primary 2.4,-1.4 User",
            "STARTOCCURRENCE 1 2.4,-1.4 17-Feb-2003 16:54:40",
            "/STARTOCCURRENCE primary User",
            "STARTWORKFLOW workflow-template workflow-uid 3 Bone",
            "CHECKPOINT 2.4,-1.4 17-Feb-2003 16:55:00",
            "CANCELOCCURRENCE 17-Feb-2003 16:56:00",
            "CANCELTRANSECT 17-Feb-2003 16:57:00",
        ])
        result = parse_log(9, payload)
        self.assertEqual(result.format, "Bones line log")
        self.assertEqual(result.transects[0]["uid"], 4354068)
        self.assertEqual(result.transects[0]["state"], "cancelled")
        self.assertEqual(result.occurrences[0]["number"], 1)
        self.assertEqual(result.occurrences[0]["state"], "cancelled")
        self.assertEqual(result.instances[0]["workflow_uid"], "workflow-uid")
        self.assertEqual(result.instances[0]["number"], 3)
        self.assertEqual(len(result.track_points), 1)
        self.assertEqual(result.track_points[0]["event"], "CHECKPOINT")
        self.assertEqual(result.track_points[0]["lat"], 2.4)
        self.assertEqual(result.track_points[0]["long"], -1.4)
        self.assertEqual(result.track_points[0]["user"], "User")
        self.assertEqual(result.instances[0]["response_count"], 0)

    def test_gps_expectation_is_explicit(self):
        self.assertEqual(gps_status(2001, 0, 0, 2010), "GPS_NOT_EXPECTED_EARLY_MANUAL")
        self.assertEqual(gps_status(2020, 0, 0, 2010), "GPS_MISSING")
        self.assertEqual(gps_status(2020, 1, 0, 2010), "GPS_PRESENT")
        self.assertEqual(gps_status(2020, 0, 0, None), "GPS_EXPECTATION_UNKNOWN")

    def test_empty_line_log_occurrence_is_identified_without_hiding_json_records(self):
        payload = "\n".join([
            "STARTTRANSECT template-guid 17 2.4,-1.4 01-Jan-2025 10:00:00 1",
            "/STARTTRANSECT primary 2.4,-1.4 User",
            "STARTOCCURRENCE 5 2.4,-1.4 01-Jan-2025 10:01:00",
            "/STARTOCCURRENCE secondary Fire",
            "ENDOCCURRENCE OUT",
            "/ENDOCCURRENCE 01-Jan-2025 10:01:02",
        ])
        occurrence = parse_log(1, payload).occurrences[0]

        self.assertTrue(is_empty_logged_occurrence(occurrence, 0))
        self.assertFalse(is_empty_logged_occurrence(occurrence, 1))
        self.assertFalse(is_empty_logged_occurrence({}, 0))

    def test_occurrence_omission_is_scoped_and_keeps_valid_merge(self):
        complete = {
            "transect_uid": 17, "evidence_count": 1,
            "parent_transect": {"state": "completed"},
        }
        self.assertFalse(omit_logged_occurrence(complete, 0, {17}))
        cancelled = {**complete, "parent_transect": {"state": "cancelled"}}
        self.assertTrue(omit_logged_occurrence(cancelled, 1, set()))
        self.assertTrue(omit_logged_occurrence(complete, 1, set(), deleted=object()))
        self.assertTrue(omit_logged_occurrence(complete, 1, set(), parent_deleted=object()))
        self.assertFalse(omit_logged_occurrence(
            complete, 1, set(), parent_deleted=object(), merged_current=object(),
        ))
        self.assertFalse(logged_parent_transect_cancelled(complete, {17}))
        self.assertTrue(logged_parent_transect_cancelled(cancelled, set()))


class ReconciliationWorkbookTests(SimpleTestCase):
    def test_pause_and_resume_are_not_recovery_events(self):
        self.assertTrue(is_recoverable_track_event("CHECKPOINT"))
        self.assertTrue(is_recoverable_track_event("TURNAROUND"))
        self.assertFalse(is_recoverable_track_event("PAUSETRANSECT"))
        self.assertFalse(is_recoverable_track_event("RESUMETRANSECT"))

    def test_invalid_coordinates_are_not_recovery_evidence(self):
        self.assertTrue(has_valid_track_coordinates(-0.04, 36.87))
        self.assertFalse(has_valid_track_coordinates(0, 0))
        self.assertFalse(has_valid_track_coordinates(None, 36.87))
        self.assertFalse(has_valid_track_coordinates(91, 36.87))

    def test_legacy_track_key_rounds_sql_precision_and_keeps_device(self):
        briana = legacy_track_key(
            17, "Briana", datetime(2025, 1, 1, 10, 39, 48),
            -0.0407316667, 36.8723816667, "CHECKPOINT",
        )
        stored = legacy_track_key(
            17, "Briana", datetime(2025, 1, 1, 10, 40),
            -0.040732, 36.872382, "CHECKPOINT",
        )
        fire = legacy_track_key(
            17, "Fire", datetime(2025, 1, 1, 10, 39, 48),
            -0.0407316667, 36.8723816667, "CHECKPOINT",
        )

        self.assertEqual(briana, stored)
        self.assertNotEqual(briana, fire)

    def test_summary_counts_status_columns_not_identifiers(self):
        rows = {
            "Transects": [[1, 17, None, "T", 2025, "MISSING"]],
            "Occurrences": [[1, 17, None, 2, None, "DELETED_CONFIRMED"]],
            "Instances": [[1, 17, None, 2, 3, "", [], "HISTORICAL_ONLY"]],
        }
        summary = dict(summary_rows(rows))
        self.assertEqual(summary["Status: MISSING"], 1)
        self.assertEqual(summary["Status: DELETED_CONFIRMED"], 1)
        self.assertEqual(summary["Status: HISTORICAL_ONLY"], 1)

    def test_deleted_entity_status_is_never_presentation_work(self):
        rows = {
            "Transects": [[1, 17, None, "T", 2025, "DELETED_CONFIRMED"]],
            "Occurrences": [], "Instances": [], "GPS": [],
            "Critical findings": [],
            "Recovery candidates": [["Occurrence", "Keep", "READY_FOR_IMPORT"]],
        }
        filter_reconciliation_rows(rows, {"DELETED_CONFIRMED"})
        self.assertEqual(rows["Transects"], [])
        self.assertEqual(rows["Recovery candidates"], [["Occurrence", "Keep", "READY_FOR_IMPORT"]])

    def test_workbook_has_review_sheets_and_headers(self):
        with tempfile.TemporaryDirectory() as folder:
            target = write_workbook(Path(folder) / "report.xlsx", {"Summary": [["Logs", 2]]})
            workbook = load_workbook(target, read_only=True)
            self.assertEqual(workbook.sheetnames, list(SHEETS))
            self.assertEqual(workbook["Summary"]["A1"].value, "Metric")
            self.assertEqual(workbook["Summary"]["A2"].value, "Logs")
            workbook.close()

    def test_workbook_can_omit_unselected_sheets_but_keeps_summary(self):
        workbook = load_workbook(BytesIO(workbook_bytes({}, ["GPS"])), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Summary", "GPS"])
        workbook.close()


class ParentTransectDeletionTests(SimpleTestCase):
    def test_cancelled_deleted_and_empty_instances_are_omitted(self):
        complete = {
            "transect_uid": 17, "state": "completed", "response_count": 1,
        }
        self.assertTrue(omit_logged_instance(complete, {17}))
        self.assertFalse(omit_logged_instance(
            {**complete, "parent_transect": {"state": "completed"}}, {17},
        ))
        self.assertTrue(omit_logged_instance(
            {**complete, "parent_transect": {"state": "cancelled"}}, set(),
        ))
        self.assertFalse(omit_logged_instance(
            {**complete, "parent_transect": {"state": "cancelled"}}, set(),
            current_matches=[object()],
        ))
        self.assertTrue(omit_logged_instance(complete, set(), deleted=object()))
        self.assertTrue(omit_logged_instance(complete, set(), parent_deleted=object()))
        self.assertTrue(omit_logged_instance(
            {"transect_uid": 17, "state": "started", "response_count": 0}, set(),
        ))
        self.assertFalse(omit_logged_instance(complete, set()))

    def test_all_workflow_uids_identify_occurrence_merged_to_another_transect(self):
        entry = {"transect_uid": 17, "number": 3}
        logged = [{"workflow_uid": "wf-1"}, {"workflow_uid": "wf-2"}]
        workflows = {
            "wf-1": SimpleNamespace(occurrence_id=91),
            "wf-2": SimpleNamespace(occurrence_id=91),
        }
        current = SimpleNamespace(id=91, transect_id=23, occurrence_number=3)

        self.assertIs(
            merged_occurrence_candidate(entry, logged, workflows, {91: current}),
            current,
        )
        self.assertIsNone(merged_occurrence_candidate(
            entry, logged, {"wf-1": workflows["wf-1"]}, {91: current},
        ))

    def test_exact_workflow_uid_recognizes_same_transect_number_merge(self):
        entry = {"transect_uid": 17, "number": 11}
        logged = [{"workflow_uid": "wf-1"}]
        workflow = SimpleNamespace(occurrence_id=91)
        current = SimpleNamespace(id=91, transect_id=17, occurrence_number=10)

        self.assertIs(
            merged_occurrence_candidate(
                entry, logged, {"wf-1": workflow}, {91: current},
            ),
            current,
        )

    def test_cross_device_occurrence_alias_uses_unique_time_and_coordinates(self):
        current = SimpleNamespace(id=616, transect_id=3967792, occurrence_number=4)
        fire = {
            "log_id": 12, "transect_uid": 3967792, "number": 5,
            "start_time": datetime(2018, 8, 18, 10, 59, 42),
            "lat": 0.0779399995, "long": 36.917475, "user": "Fire",
            "state": "completed", "evidence_count": 1,
        }
        briana = {
            "log_id": 6, "transect_uid": 3967792, "number": 4,
            "start_time": datetime(2018, 8, 18, 11, 0, 55),
            "lat": 0.0779, "long": 36.9174899328, "user": "Briana",
            "state": "completed", "evidence_count": 13,
        }
        candidates = defaultdict(list)
        candidates[(3967792, 4)].append(current)

        self.assertIs(
            cross_device_occurrence_candidate(fire, [fire, briana], candidates),
            current,
        )
        self.assertFalse(omit_logged_occurrence(
            {**fire, "evidence_count": 0}, 0, set(), alias_current=current,
        ))

    def test_cross_device_occurrence_alias_rejects_ambiguous_matches(self):
        fire = {
            "log_id": 12, "transect_uid": 7, "number": 5,
            "start_time": datetime(2018, 8, 18, 10, 0),
            "lat": 1.0, "long": 2.0, "user": "Fire",
            "state": "completed", "evidence_count": 0,
        }
        peers = [
            {
                **fire, "log_id": index, "number": number, "user": "Briana",
                "evidence_count": 1,
            }
            for index, number in ((1, 3), (2, 4))
        ]
        candidates = defaultdict(list)
        candidates[(7, 3)].append(SimpleNamespace(id=3))
        candidates[(7, 4)].append(SimpleNamespace(id=4))

        self.assertIsNone(
            cross_device_occurrence_candidate(fire, [fire, *peers], candidates)
        )

    def test_parent_deletion_confirms_logged_occurrence_and_instance_deletions(self):
        parsed = SimpleNamespace(
            error="", log_id=1, track_points=[],
            transects=[{"log_id": 1, "uid": 17, "name": "Test", "template": "tt", "start_time": datetime(2025, 1, 1), "state": "completed", "source": "line 1"}],
            occurrences=[{"log_id": 1, "transect_uid": 17, "id": None, "number": 1, "start_time": datetime(2025, 1, 1, 0, 1), "lat": 1, "long": 2, "state": "completed", "source": "line 2"}],
            instances=[{"log_id": 1, "transect_uid": 17, "occurrence_id": None, "occurrence_number": 1, "number": 1, "workflow_uid": "wf", "template": "tw", "state": "completed", "response_count": 1, "parent_occurrence": {"user": "Tester"}, "source": "line 3"}],
        )
        model_names = (
            "CompletedTransect", "CompletedOccurrence", "CompletedWorkflow",
            "CompletedResponse", "CompletedTransectTrack", "TransectDataLog",
            "TemplateTransect", "TemplateWorkflow", "OccurrenceDeletion",
            "InstanceDeletion",
        )
        deletion = SimpleNamespace(
            transect_uid=17, deleted_at=datetime(2025, 1, 2),
            reason="Test transect", pk="audit-1",
        )
        with ExitStack() as stack:
            for name in model_names:
                stack.enter_context(patch(f"bones.management.commands.reconcile_data_logs.{name}.objects"))
            for name in ("CompletedTransect", "CompletedOccurrence", "CompletedWorkflow"):
                stack.enter_context(patch(f"bones.management.commands.reconcile_data_logs.{name}.history"))
            transect_deletions = stack.enter_context(
                patch("bones.management.commands.reconcile_data_logs.TransectDeletion.objects")
            )
            transect_deletions.all.return_value = [deletion]
            rows = Command()._reconcile(
                [parsed], [{"id": 1, "upload_date": None, "uploaded_by": "Tester"}],
                {"from_year": None, "to_year": None, "gps_required_from_year": None},
            )

        self.assertEqual(rows["Occurrences"], [])
        self.assertEqual(rows["Instances"], [])
        recovery = {(row[0], row[2]) for row in rows["Recovery candidates"]}
        self.assertNotIn(("Occurrence", "INTENTIONALLY_DELETED"), recovery)
        self.assertNotIn(("Instance", "INTENTIONALLY_DELETED"), recovery)


class ReconciliationReportUITests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_defaults_are_exception_focused(self):
        form = DataReconciliationReportForm()
        statuses = set(form.fields["statuses"].initial)
        self.assertIn("MISSING", statuses)
        self.assertNotIn("CURRENT_EXACT", statuses)
        self.assertNotIn("LOG_CANCELLED", statuses)
        self.assertNotIn("GPS_PRESENT", statuses)
        self.assertNotIn("DELETED_CONFIRMED", dict(form.fields["statuses"].choices))
        recovery = set(form.fields["recovery_statuses"].initial)
        self.assertIn("READY_FOR_IMPORT", recovery)
        self.assertNotIn("INTENTIONALLY_DELETED", recovery)
        self.assertNotIn("INTENTIONALLY_DELETED", dict(form.fields["recovery_statuses"].choices))
        self.assertNotIn("CANCELLED_IN_LOG", dict(form.fields["recovery_statuses"].choices))

    def test_log_selector_defers_large_payload_contents(self):
        form = DataReconciliationReportForm()
        deferred, is_deferred = form.fields["logs"].queryset.query.deferred_loading
        self.assertTrue(is_deferred)
        self.assertIn("contents", deferred)

    def test_year_range_must_be_chronological(self):
        form = DataReconciliationReportForm(data={
            "from_year": 2025, "to_year": 2020,
            "contents": ["Critical findings"], "statuses": ["MISSING"],
        })
        self.assertFalse(form.is_valid())
        self.assertIn("to_year", form.errors)

    def test_view_uses_dedicated_permission(self):
        self.assertEqual(
            DataReconciliationReportView.permission_required,
            "bones.run_data_reconciliation_report",
        )

    def test_route_and_reports_navigation_are_available(self):
        url = reverse("bones:reports:data_reconciliation")
        self.assertEqual(url, "/reports/data-reconciliation/")
        reports = next(section for section in navigation_context(None)["navigation_sections"] if section["label"] == "Reports")
        link = next(child for child in reports["children"] if child["label"] == "Data Reconciliation")
        self.assertEqual(link["url"], url)

    def test_template_is_semantic_responsive_and_accessible(self):
        template = (Path(__file__).resolve().parents[1] / "templates" / "bones" / "reports" / "data_reconciliation.html").read_text(encoding="utf-8")
        self.assertIn("<header", template)
        self.assertIn("<section", template)
        self.assertIn("w3-row-padding", template)
        self.assertIn('aria-label="Data reconciliation report filters"', template)
        self.assertIn("fa-file-excel", template)

    @patch("bones.views.reports.workbook_bytes", return_value=b"xlsx")
    @patch("bones.views.reports.collect_reconciliation_rows")
    @patch("bones.views.reports.DataReconciliationReportForm")
    def test_valid_export_returns_timestamped_workbook(self, form_class, collect_rows, build_workbook):
        form = MagicMock()
        form.is_valid.return_value = True
        form.cleaned_data = {
            "from_year": None, "to_year": None, "logs": [],
            "gps_required_from_year": None, "statuses": ["MISSING"],
            "recovery_statuses": ["READY_FOR_IMPORT"],
            "contents": ["Critical findings", "Methodology"],
        }
        form_class.return_value = form
        collect_rows.return_value = ({name: [] for name in SHEETS}, 37)
        request = self.factory.get("/reports/data-reconciliation/", {"export": "xlsx"})
        request.user = SimpleNamespace(
            is_authenticated=True, has_perms=lambda permissions: True,
            get_full_name=lambda: "Report User", get_username=lambda: "report-user",
        )
        response = DataReconciliationReportView.as_view()(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b"xlsx")
        self.assertIn("data-reconciliation-", response["Content-Disposition"])
        build_workbook.assert_called_once()
