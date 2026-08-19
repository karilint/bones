from io import BytesIO
from html import unescape
from types import SimpleNamespace
from unittest.mock import patch

from django.test import SimpleTestCase
from django.template.loader import render_to_string
from django.urls import reverse
from openpyxl import load_workbook

from ..reports.weathering import (
    CALCULATION_RULES, WeatheringReportResult, _has_carnivore_damage,
    build_weathering_report,
)
from ..views.reports import export_weathering_workbook


class WeatheringCalculationTests(SimpleTestCase):
    def test_carnivore_damage_supports_current_and_legacy_answers(self):
        self.assertTrue(_has_carnivore_damage({"carnivore damage": "2"}))
        self.assertTrue(_has_carnivore_damage({"carn damage 3 level": "1"}))
        self.assertTrue(_has_carnivore_damage({"carn damage 2 portion": "HUM SH"}))
        self.assertFalse(_has_carnivore_damage({"carnivore damage": "0"}))

    @patch("bones.reports.weathering.rows_in_batches")
    @patch("bones.reports.weathering.build_report")
    def test_matrix_calculates_habitat_and_all_from_raw_counts(
        self, build_mni, batched_rows
    ):
        transects = [
            SimpleNamespace(pk=1, distance_km=10),
            SimpleNamespace(pk=2, distance_km=20),
        ]
        mni_result = SimpleNamespace(
            transect_metadata={
                1: {"habitat": "open grassland"},
                2: {"habitat": "shrubs closed"},
            },
            transect_rows=[
                {"transect_id": 1, "mni": 3},
                {"transect_id": 2, "mni": 2},
            ],
            occurrence_metadata={
                10: {"transect_id": 1}, 20: {"transect_id": 2}
            },
        )
        build_mni.return_value = (mni_result, transects, [])
        workflows = [
            SimpleNamespace(pk="a", occurrence=SimpleNamespace(transect_id=1)),
            SimpleNamespace(pk="b", occurrence=SimpleNamespace(transect_id=1)),
            SimpleNamespace(pk="c", occurrence=SimpleNamespace(transect_id=2)),
        ]
        responses = [
            {"workflow_id": "a", "question_text": "Weathering class", "response": "0"},
            {"workflow_id": "a", "question_text": "Buried?", "response": "Yes"},
            {"workflow_id": "b", "question_text": "Weathering class", "response": "0-1"},
            {"workflow_id": "b", "question_text": "Carn Damage 1 Portion", "response": "HUM SH"},
            {"workflow_id": "c", "question_text": "Weathering class", "response": "unknown"},
        ]
        batched_rows.side_effect = [workflows, responses]

        result = build_weathering_report({})
        rows = {row["label"]: row["values"] for row in result.rows}

        self.assertEqual(result.habitats, ["open grassland", "shrubs closed", "All"])
        self.assertEqual(rows["# transects"], [1, 1, 2])
        self.assertEqual(rows["km² covered"], [0.5, 1.0, 1.5])
        self.assertEqual(rows["Count of occurrences"], [1, 1, 2])
        self.assertEqual(rows["MNI"], [3, 2, 5])
        self.assertEqual(rows["NISP (# of bones)"], [2, 1, 3])
        self.assertEqual(rows["% bones W0"], [50, 0, 100 / 3])
        self.assertEqual(rows["% bones W0-1"], [50, 0, 100 / 3])
        self.assertEqual(rows["% bones part buried"], [50, 0, 100 / 3])
        self.assertEqual(rows["% bones with any carnivore damage"], [50, 0, 100 / 3])
        self.assertIn("missing or unrecognised", result.warnings[0])
        self.assertEqual(build_mni.call_args.args[0], {})


class WeatheringPresentationTests(SimpleTestCase):
    def test_route_is_available(self):
        self.assertEqual(reverse("bones:reports:weathering"), "/reports/weathering/")

    def test_export_contains_matrix_and_exact_calculation_rules(self):
        result = WeatheringReportResult(
            habitats=["open grassland", "All"],
            rows=[{"label": "# transects", "kind": "integer", "values": [2, 2]}],
            warnings=["Example warning"],
        )
        workbook = load_workbook(BytesIO(export_weathering_workbook(
            result, "2 eligible transects", CALCULATION_RULES
        )))

        self.assertEqual(
            workbook.sheetnames,
            ["Weathering", "Calculation rules", "Data quality"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["Weathering"][1]],
            ["Measure", "open grassland", "All"],
        )
        exported_rules = [
            row[1].value
            for row in workbook["Calculation rules"].iter_rows(min_row=2)
            if isinstance(row[0].value, int)
        ]
        self.assertEqual(exported_rules, list(CALCULATION_RULES))
        self.assertEqual(workbook["Data quality"]["A2"].value, "Example warning")

    def test_screen_displays_the_same_exact_calculation_rules(self):
        result = WeatheringReportResult(
            habitats=["All"],
            rows=[{"label": "# transects", "kind": "integer", "values": [1]}],
        )
        html = render_to_string("bones/reports/weathering.html", {
            "page_title": "Weathering report", "form": [],
            "note_filter": SimpleNamespace(
                note_filter_form_rows=[], note_filter_choices=[]
            ),
            "result": result, "report_error": "", "filter_summary": "one",
            "calculation_rules": CALCULATION_RULES, "export_query": "export=xlsx",
        })
        html = unescape(html)

        self.assertIn("Exact calculation rules", html)
        for rule in CALCULATION_RULES:
            self.assertIn(rule, html)
