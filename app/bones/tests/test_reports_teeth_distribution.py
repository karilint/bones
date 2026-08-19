from html import unescape
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from ..reports.teeth_distribution import CALCULATION_RULES, TeethDistributionResult, build_teeth_distribution_report
from ..views.reports import export_teeth_distribution_workbook


class TeethDistributionTests(SimpleTestCase):
    @patch("bones.reports.teeth_distribution.rows_in_batches")
    @patch("bones.reports.teeth_distribution.build_report")
    def test_dentition_denominator_indet_alias_and_unlisted_warning(self, build_mni, batched_rows):
        build_mni.return_value = (SimpleNamespace(
            transect_metadata={1: {"habitat": "grass"}}, occurrence_metadata={10: {}}
        ), [SimpleNamespace(pk=1)], [])
        workflows = [SimpleNamespace(pk=x, occurrence=SimpleNamespace(transect_id=1)) for x in "abc"]
        batched_rows.side_effect = [workflows, [
            {"workflow_id": "a", "response": "indet"},
            {"workflow_id": "b", "response": "M1"},
            {"workflow_id": "c", "response": "homodont"},
        ]]
        result = build_teeth_distribution_report({})
        rows = {row["label"]: row["values"] for row in result.rows}
        self.assertEqual(rows["number of specimens"], [3, 3])
        self.assertEqual(rows["tooth indet"], [100 / 3, 100 / 3])
        self.assertEqual(rows["M1"], [100 / 3, 100 / 3])
        self.assertIn("homodont", result.warnings[0])

    def test_route_screen_and_export_share_exact_rules(self):
        self.assertEqual(reverse("bones:reports:teeth_distribution"), "/reports/teeth-distribution/")
        result = TeethDistributionResult(["All"], [
            {"label": "number of specimens", "kind": "count", "values": [1]},
            {"label": "C", "kind": "percent", "values": [100]},
        ])
        workbook = load_workbook(BytesIO(export_teeth_distribution_workbook(result, "all", CALCULATION_RULES)))
        exported = [row[1].value for row in workbook["Calculation rules"].iter_rows(min_row=2) if isinstance(row[0].value, int)]
        self.assertEqual(exported, list(CALCULATION_RULES))
        html = unescape(render_to_string("bones/reports/teeth_distribution.html", {
            "page_title": "Teeth distribution report", "form": [],
            "note_filter": SimpleNamespace(note_filter_form_rows=[], note_filter_choices=[]),
            "result": result, "report_error": "", "filter_summary": "all",
            "calculation_rules": CALCULATION_RULES, "export_query": "export=xlsx",
        }))
        for rule in CALCULATION_RULES:
            self.assertIn(rule, html)
