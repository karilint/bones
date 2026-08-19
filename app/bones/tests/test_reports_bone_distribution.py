from html import unescape
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from ..reports.bone_distribution import (
    CALCULATION_RULES, BoneDistributionResult,
    build_bone_distribution_report,
)
from ..views.reports import export_bone_distribution_workbook


class BoneDistributionCalculationTests(SimpleTestCase):
    @patch("bones.reports.bone_distribution.rows_in_batches")
    @patch("bones.reports.bone_distribution.build_report")
    def test_counts_percentages_alias_null_and_subheaders(
        self, build_mni, batched_rows
    ):
        transects = [SimpleNamespace(pk=1)]
        build_mni.return_value = (
            SimpleNamespace(
                transect_metadata={1: {"habitat": "grass"}},
                occurrence_metadata={10: {"transect_id": 1}},
            ), transects, [],
        )
        workflows = [
            SimpleNamespace(pk="a", occurrence=SimpleNamespace(transect_id=1)),
            SimpleNamespace(pk="b", occurrence=SimpleNamespace(transect_id=1)),
            SimpleNamespace(pk="c", occurrence=SimpleNamespace(transect_id=1)),
        ]
        responses = [
            {"workflow_id": "a", "response": "cranium"},
            {"workflow_id": "b", "response": "bone nonidentifiable"},
        ]
        batched_rows.side_effect = [workflows, responses]

        result = build_bone_distribution_report({})
        rows = {row["label"]: row for row in result.rows}

        self.assertEqual(rows["number of specimens"]["values"], [3, 3])
        self.assertEqual(rows["cranium"]["values"], [100 / 3, 100 / 3])
        self.assertEqual(rows["bone non-identifiable"]["values"], [100 / 3, 100 / 3])
        self.assertEqual(rows["<null>"]["values"], [100 / 3, 100 / 3])
        self.assertEqual(rows["Skull Bones:"]["kind"], "subheader")
        self.assertEqual(rows["Facial Bones:"]["kind"], "subheader")


class BoneDistributionPresentationTests(SimpleTestCase):
    def test_route_is_available(self):
        self.assertEqual(
            reverse("bones:reports:bone_distribution"),
            "/reports/bone-distribution/",
        )

    def test_screen_and_export_preserve_subheaders_and_exact_rules(self):
        result = BoneDistributionResult(habitats=["All"], rows=[
            {"label": "number of specimens", "kind": "count", "values": [1]},
            {"label": "Skull Bones:", "kind": "subheader", "values": []},
            {"label": "cranium", "kind": "percent", "values": [100]},
        ])
        workbook = load_workbook(BytesIO(export_bone_distribution_workbook(
            result, "all eligible", CALCULATION_RULES
        )))
        self.assertIn("A3:B3", [str(value) for value in workbook["Bone distribution"].merged_cells.ranges])
        exported = [
            row[1].value
            for row in workbook["Calculation rules"].iter_rows(min_row=2)
            if isinstance(row[0].value, int)
        ]
        self.assertEqual(exported, list(CALCULATION_RULES))

        html = unescape(render_to_string("bones/reports/bone_distribution.html", {
            "page_title": "Bone distribution report", "form": [],
            "note_filter": SimpleNamespace(note_filter_form_rows=[], note_filter_choices=[]),
            "result": result, "report_error": "", "filter_summary": "all eligible",
            "calculation_rules": CALCULATION_RULES, "export_query": "export=xlsx",
        }))
        self.assertIn('scope="rowgroup"', html)
        self.assertIn("Skull Bones:", html)
        for rule in CALCULATION_RULES:
            self.assertIn(rule, html)
