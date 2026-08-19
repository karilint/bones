from html import unescape
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from ..reports.carnivore import (
    CALCULATION_RULES, CarnivoreReportResult, _damage_level,
    build_carnivore_report,
)
from ..views.reports import export_carnivore_workbook


class CarnivoreCalculationTests(SimpleTestCase):
    def test_primary_legacy_level_precedes_parent_value(self):
        self.assertEqual(_damage_level({
            "carn damage 1 level": "2", "carnivore damage": "4"
        }), (2, False))
        self.assertEqual(_damage_level({"carnivore damage": "3"}), (3, False))
        self.assertEqual(_damage_level({}), (0, False))

    @patch("bones.reports.carnivore.rows_in_batches")
    @patch("bones.reports.carnivore.build_report")
    def test_report_uses_one_damage_and_primary_mark_class_per_bone(
        self, build_mni, batched_rows
    ):
        transects = [SimpleNamespace(pk=1)]
        mni_result = SimpleNamespace(
            transect_metadata={1: {"habitat": "grass", "block": "T1"}},
            occurrence_metadata={10: {
                "transect_id": 1, "number": 1, "taxon": "Zebra",
            }},
        )
        build_mni.return_value = (mni_result, transects, [])
        workflows = [
            SimpleNamespace(pk="a", occurrence_id=10, instance_number=1,
                            occurrence=SimpleNamespace(transect_id=1)),
            SimpleNamespace(pk="b", occurrence_id=10, instance_number=2,
                            occurrence=SimpleNamespace(transect_id=1)),
        ]
        responses = [
            {"workflow_id": "a", "question_text": "Carn Damage 1 Level", "response": "2"},
            {"workflow_id": "a", "question_text": "Carn Damage 2 Level", "response": "4"},
            {"workflow_id": "a", "question_text": "Portion 1: TM Type", "response": "pit"},
            {"workflow_id": "a", "question_text": "Portion 2: TM Type", "response": "score"},
            {"workflow_id": "a", "question_text": "Portion 2: TM Number", "response": "3"},
            {"workflow_id": "a", "question_text": "Complete?", "response": "No"},
            {"workflow_id": "a", "question_text": "What element is this?", "response": "Humerus"},
            {"workflow_id": "a", "question_text": "Long Bone Portion", "response": "00111"},
            {"workflow_id": "b", "question_text": "Carnivore damage", "response": "1"},
            {"workflow_id": "b", "question_text": "Portion 1: TM Type", "response": "puncture"},
            {"workflow_id": "b", "question_text": "Complete?", "response": "Yes"},
        ]
        batched_rows.side_effect = [workflows, responses]

        result = build_carnivore_report({})
        rows = {row["label"]: row["values"] for row in result.rows}

        self.assertEqual(result.habitats, ["grass", "All"])
        self.assertEqual(rows["% bones C1"], [50, 50])
        self.assertEqual(rows["% bones C2"], [50, 50])
        self.assertEqual(rows["% bones C4"], [0, 0])
        self.assertEqual(rows["% pits"], [50, 50])
        self.assertEqual(rows["% punctures"], [50, 50])
        self.assertEqual(rows["% scores"], [0, 0])
        self.assertEqual(rows["Number of specimens"], [2, 2])
        self.assertEqual(rows["% complete specimens"], [50, 50])
        self.assertEqual(len(result.specimen_rows), 2)
        self.assertEqual(len(result.damage_rows), 3)
        self.assertIn("Long Bone Portion: 00111", result.specimen_rows[0])
        self.assertEqual(result.damage_rows[1][-2:], ["score", "3"])


class CarnivorePresentationTests(SimpleTestCase):
    def test_route_is_available(self):
        self.assertEqual(
            reverse("bones:reports:carnivore"), "/reports/carnivore-damage/"
        )

    def test_screen_and_export_contain_identical_exact_rules(self):
        result = CarnivoreReportResult(
            habitats=["All"],
            rows=[{"label": "% bones C0", "kind": "percent", "values": [100]}],
        )
        workbook = load_workbook(BytesIO(export_carnivore_workbook(
            result, "all eligible", CALCULATION_RULES
        )))
        exported = [
            row[1].value
            for row in workbook["Calculation rules"].iter_rows(min_row=2)
            if isinstance(row[0].value, int)
        ]
        self.assertEqual(exported, list(CALCULATION_RULES))
        self.assertIn("Specimen preservation", workbook.sheetnames)
        self.assertIn("Damage observations", workbook.sheetnames)

        html = unescape(render_to_string("bones/reports/carnivore.html", {
            "page_title": "Carnivore damage report", "form": [],
            "note_filter": SimpleNamespace(
                note_filter_form_rows=[], note_filter_choices=[]
            ),
            "result": result, "report_error": "",
            "filter_summary": "all eligible",
            "calculation_rules": CALCULATION_RULES, "export_query": "export=xlsx",
        }))
        for rule in CALCULATION_RULES:
            self.assertIn(rule, html)
