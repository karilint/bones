from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch
from pathlib import Path
from django.contrib import admin
from django.test import RequestFactory, SimpleTestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from ..admin_general import (BonesHistoryAdmin, DataLogAdmin, OccurrenceInfoAdmin, ResponseAdmin,
                             WorkflowAdmin, WorkflowDeletionForm)
from ..models import CompletedOccurrence, CompletedOccurrenceInfo, CompletedResponse, CompletedTransect, CompletedWorkflow, DataLogFile
from ..occurrence_info_imports import COLUMN_GUIDE, HEADERS, template_workbook, validate_workbook

class GeneralAdminTests(SimpleTestCase):
    def setUp(self):
        self.factory=RequestFactory()
        self.user=SimpleNamespace(has_perm=lambda permission: True)

    def test_main_operational_models_are_registered(self):
        for model in (CompletedTransect,CompletedOccurrence,CompletedWorkflow,CompletedResponse):
            with self.subTest(model=model.__name__): self.assertIn(model,admin.site._registry)

    @patch("bones.admin_general.ReadOnlyAdmin.get_queryset")
    def test_data_log_admin_defers_contents(self, get_queryset):
        base_queryset = MagicMock()
        get_queryset.return_value = base_queryset
        request = self.factory.get("/admin/bones/datalogfile/")
        request.user = self.user

        model_admin = admin.site._registry[DataLogFile]
        self.assertIsInstance(model_admin, DataLogAdmin)
        model_admin.get_queryset(request)

        base_queryset.defer.assert_called_once_with("contents")

    def test_answers_use_audited_admin(self):
        model_admin=admin.site._registry[CompletedResponse]
        self.assertIsInstance(model_admin,ResponseAdmin)
        self.assertIsInstance(model_admin,BonesHistoryAdmin)

    def test_answer_hierarchy_is_readonly(self):
        model_admin=admin.site._registry[CompletedResponse]
        for field in ("id","occurrence","workflow","question","question_number","question_text"):
            self.assertIn(field,model_admin.readonly_fields)
        for field in ("response_code","response","skipped"):
            self.assertIn(field,model_admin.fields)

    def test_edit_reason_is_required_for_existing_answer(self):
        model_admin=admin.site._registry[CompletedResponse]
        request=self.factory.get("/admin/bones/completedresponse/1/change/"); request.user=self.user
        instance=CompletedResponse(id=1,occurrence_id=2,workflow_id="w",question_id=None,question_text="Free text",skipped=False)
        form_class=model_admin.get_form(request,instance,change=True)
        form=form_class(instance=instance)
        self.assertTrue(form.fields["edit_reason"].required)
        self.assertIn("record_version",form.fields)
        self.assertIn("answer_option",form.fields)
        self.assertIsInstance(form.fields["answer_option"].widget, __import__("django").forms.HiddenInput)

    def test_operational_admins_disable_delete_and_bulk_delete(self):
        request=self.factory.get("/admin/"); request.user=self.user
        for model in (CompletedTransect,CompletedOccurrence,CompletedWorkflow,CompletedResponse):
            model_admin=admin.site._registry[model]
            with self.subTest(model=model.__name__):
                self.assertFalse(model_admin.has_delete_permission(request))
                self.assertNotIn("delete_selected",model_admin.get_actions(request))

    def test_workflow_removal_requires_reason_and_confirmation(self):
        self.assertFalse(WorkflowDeletionForm({"reason":"Correction"}).is_valid())
        self.assertFalse(WorkflowDeletionForm({"confirm":True}).is_valid())
        self.assertTrue(WorkflowDeletionForm({"reason":"Duplicate workflow","confirm":True}).is_valid())

    def test_workflow_admin_has_dedicated_removal_route(self):
        model_admin=admin.site._registry[CompletedWorkflow]
        self.assertIsInstance(model_admin,WorkflowAdmin)
        self.assertEqual(
            reverse("admin:bones_completedworkflow_remove",args=("workflow-1",)),
            "/admin/bones/completedworkflow/workflow-1/remove/",
        )

    def test_workflow_removal_template_documents_safeguards(self):
        template=(Path(__file__).resolve().parents[1]/"templates"/"admin"/"bones"/"completedworkflow"/"remove_workflow.html").read_text(encoding="utf-8")
        self.assertIn("last workflow in the instance",template)
        self.assertIn("Responses to delete",template)
        self.assertIn("Instance images",template)

    def test_occurrence_notes_use_audited_admin(self):
        model_admin=admin.site._registry[CompletedOccurrenceInfo]
        self.assertIsInstance(model_admin,OccurrenceInfoAdmin)
        self.assertIsInstance(model_admin,BonesHistoryAdmin)
        for field in ("id","occurrence","pre_or_post","question_text","response_data_type"):
            self.assertIn(field,model_admin.readonly_fields)
        for field in ("answer_option","response_code","response","edit_reason","record_version"):
            self.assertIn(field,model_admin.fields)

    def test_occurrence_info_admin_has_bulk_update_routes(self):
        self.assertEqual(
            reverse("admin:bones_completedoccurrenceinfo_bulk_update"),
            "/admin/bones/completedoccurrenceinfo/bulk-update/",
        )
        self.assertEqual(
            reverse("admin:bones_completedoccurrenceinfo_import_template"),
            "/admin/bones/completedoccurrenceinfo/bulk-update/template/",
        )

    def test_occurrence_info_import_template_has_contract_and_instructions(self):
        workbook=load_workbook(BytesIO(template_workbook()))
        self.assertEqual(tuple(cell.value for cell in workbook["Updates"][1]),HEADERS)
        self.assertIn("Instructions",workbook.sheetnames)
        self.assertEqual(workbook["Updates"]["C2"].value,"Post")
        self.assertIn("not the template name",workbook["Updates"]["A1"].comment.text)
        self.assertIn("not the occurrence database key",workbook["Updates"]["B1"].comment.text)
        self.assertIn("not a question ID",workbook["Updates"]["D1"].comment.text)
        instructions=workbook["Instructions"]
        self.assertIn("human-readable values",instructions["A2"].value)
        self.assertEqual(tuple(cell.value for cell in instructions[8]),("Column","Required","Human-readable value expected","Example","Validation"))
        self.assertEqual(tuple(instructions.cell(row=9,column=index).value for index in range(1,6)),COLUMN_GUIDE[0])

    def test_occurrence_info_import_rejects_long_comment_before_database_lookup(self):
        workbook=Workbook(); sheet=workbook.active; sheet.title="Updates"
        sheet.append(HEADERS); sheet.append(("1",5,"Post","Taxon Guess?","waterbuck","x"*101))
        content=BytesIO(); workbook.save(content)
        result=validate_workbook(content.getvalue())
        self.assertEqual(result[0]["status"],"error")
        self.assertIn("100",result[0]["message"])

    def test_occurrence_info_import_rejects_wrong_headers(self):
        workbook=Workbook(); workbook.active.title="Updates"; workbook.active.append(("wrong",))
        content=BytesIO(); workbook.save(content)
        with self.assertRaisesMessage(ValueError,"headings"):
            validate_workbook(content.getvalue())

    @patch("bones.occurrence_info_imports.CompletedOccurrenceInfo.objects.select_related")
    def test_occurrence_info_import_matches_completed_transect_name(self,select_related):
        target=SimpleNamespace(
            pk=12,response_data_type=None,response_code="41",response="impala",occurrence_id=7,
            occurrence=SimpleNamespace(transect_id=99,transect=SimpleNamespace(name="133")),
        )
        queryset=MagicMock(); filtered=MagicMock(); filtered.__getitem__.return_value=[target]
        queryset.filter.return_value=filtered; select_related.return_value=queryset
        workbook=Workbook(); sheet=workbook.active; sheet.title="Updates"
        sheet.append(HEADERS); sheet.append(("133",2,"Post","Taxon Guess?","waterbuck","Corrected taxon"))
        content=BytesIO(); workbook.save(content)
        result=validate_workbook(content.getvalue())
        self.assertEqual(result[0]["status"],"ready")
        kwargs=queryset.filter.call_args.kwargs
        self.assertEqual(kwargs["occurrence__transect__name__iexact"],"133")
        self.assertNotIn("occurrence__transect__transect_template__name__iexact",kwargs)

    def test_occurrence_note_without_options_uses_free_text_fields(self):
        model_admin=admin.site._registry[CompletedOccurrenceInfo]
        request=self.factory.get("/admin/bones/completedoccurrenceinfo/1/change/"); request.user=self.user
        instance=CompletedOccurrenceInfo(id=1,occurrence_id=2,pre_or_post="Post",question_text="Taxon Guess?",response_data_type=None)
        form=model_admin.get_form(request,instance,change=True)(instance=instance)
        self.assertTrue(form.fields["edit_reason"].required)
        self.assertIsInstance(form.fields["answer_option"].widget,__import__("django").forms.HiddenInput)
        self.assertNotIsInstance(form.fields["response"].widget,__import__("django").forms.HiddenInput)

    @patch("bones.admin_general.DataTypeOption.objects.filter")
    def test_occurrence_note_with_options_uses_dropdown(self,filter_options):
        option=SimpleNamespace(code="lion",text="Lion")
        options=MagicMock(); options.order_by.return_value=[option]; filter_options.return_value=options
        model_admin=admin.site._registry[CompletedOccurrenceInfo]
        request=self.factory.get("/admin/bones/completedoccurrenceinfo/1/change/"); request.user=self.user
        instance=CompletedOccurrenceInfo(id=1,occurrence_id=2,pre_or_post="Post",question_text="Taxon Guess?",response_data_type="taxon-type",response_code="lion",response="Lion")
        form_class=model_admin.get_form(request,instance,change=True)
        form=form_class(data={"answer_option":"lion","edit_reason":"Corrected identification","record_version":__import__("bones.admin_general",fromlist=["signature"]).signature(instance)},instance=instance)
        self.assertIsInstance(form.fields["answer_option"].widget,__import__("django").forms.Select)
        self.assertIsInstance(form.fields["response_code"].widget,__import__("django").forms.HiddenInput)
        self.assertTrue(form.is_valid(),form.errors)
        self.assertEqual(form.cleaned_data["response_code"],"lion")
        self.assertEqual(form.cleaned_data["response"],"Lion")
        filter_options.assert_called_once_with(data_type_id="taxon-type")
