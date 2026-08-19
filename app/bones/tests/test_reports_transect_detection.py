from datetime import datetime, timedelta
from html import unescape
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import SimpleTestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from ..reports.transect_detection import CALCULATION_RULES, TransectDetectionResult, build_transect_detection_report
from ..forms_reports import TransectDetectionReportForm
from ..views.reports import export_transect_detection_workbook


class TransectDetectionCalculationTests(SimpleTestCase):
    @patch("bones.reports.transect_detection.rows_in_batches")
    @patch("bones.reports.transect_detection.build_report")
    def test_builds_requested_transect_and_occurrence_fields(self, build_mni, batched):
        start = datetime(2026, 1, 2, 8, 0)
        transect = SimpleNamespace(
            pk=1, name="T1", transect_template=SimpleNamespace(name="Template 1"),
            start_time=start, end_time=start + timedelta(minutes=100),
            paused_for_minutes=10, state="Completed", lat_from=1, long_from=2,
            lat_turn=3, long_turn=4, lat_to=5, long_to=6, distance_km=2.5,
        )
        mni = SimpleNamespace(
            transect_metadata={1: {"habitat": "grass", "reserve": "Yes"}},
            occurrence_metadata={
                10: {"taxon": "zebra"}, 11: {"taxon": "Unknown taxon"}
            },
            transect_rows=[{"transect_id": 1, "mni": 3}],
        )
        build_mni.return_value = (mni, [transect], [])
        occurrence = SimpleNamespace(
            pk=10, transect_id=1, occurrence_number=2, lat=7, long=8,
            recording_start_time=start + timedelta(minutes=20),
        )
        excluded_occurrence = SimpleNamespace(
            pk=11, transect_id=1, occurrence_number=3, lat=9, long=10,
            recording_start_time=start + timedelta(minutes=30),
        )
        info = [
            {"occurrence_id": 10, "pre_or_post": "Pre", "question_text": "Size class", "response": "2"},
            {"occurrence_id": 10, "pre_or_post": "Post", "question_text": "Size class", "response": "3"},
            {"occurrence_id": 10, "pre_or_post": "Pre", "question_text": "Distance spotted", "response": "5"},
        ]
        workflows = [
            SimpleNamespace(occurrence_id=10, template_workflow=SimpleNamespace(name="Bone")),
            SimpleNamespace(occurrence_id=10, template_workflow=SimpleNamespace(name="Dentition")),
        ]
        batched.side_effect = [[occurrence, excluded_occurrence], info, workflows]

        result = build_transect_detection_report({"excluded_taxa": ["Unknown taxon"]})
        self.assertFalse(build_mni.call_args.kwargs["apply_population_rules"])
        summary = dict(zip(result.transect_headers, result.transect_rows[0]))
        detail = dict(zip(result.occurrence_headers, result.occurrence_rows[0]))
        self.assertEqual(summary["Transect name"], "T1")
        self.assertEqual(summary["Search width (m, assumed)"], 50)
        self.assertNotIn("Completed?", summary)
        self.assertNotIn("Gross duration (minutes)", summary)
        self.assertNotIn("New Transect?", summary)
        self.assertEqual(summary["Occurrence count"], 2)
        self.assertEqual(summary["Bone NISP"], 1)
        self.assertEqual(summary["Dentition specimen count"], 1)
        self.assertEqual(summary["Calculated MNI"], 3)
        self.assertEqual(detail["Size class"], "3")
        self.assertEqual(detail["Taxon"], "zebra")
        self.assertEqual(detail["Scatter diameter"], 0)
        self.assertEqual(len(result.occurrence_rows), 1)


class TransectDetectionPresentationTests(SimpleTestCase):
    @patch("bones.forms_reports.MNIReportForm.__init__", return_value=None)
    def test_report_has_no_default_taxon_exclusions(self, _form_init):
        form = object.__new__(TransectDetectionReportForm)
        form.initial = {"excluded_taxa": ["Unknown taxon"]}
        TransectDetectionReportForm.__init__(form)
        self.assertEqual(form.initial["excluded_taxa"], [])

    def test_route_screen_and_export(self):
        self.assertEqual(reverse("bones:reports:transect_detection"), "/reports/transect-detection/")
        aware_time = timezone.make_aware(datetime(2026, 1, 2, 8, 30))
        result = TransectDetectionResult(
            ["Transect UID"], [[1]],
            ["Occurrence ID", "Recording time"], [[10, aware_time]],
        )
        workbook = load_workbook(BytesIO(export_transect_detection_workbook(result, "all", CALCULATION_RULES)))
        self.assertEqual(workbook.sheetnames, ["Transect summary", "Occurrence detection", "Calculation rules"])
        exported_time = workbook["Occurrence detection"]["B2"].value
        self.assertIsNone(exported_time.tzinfo)
        html = unescape(render_to_string("bones/reports/transect_detection.html", {
            "page_title": "Transect detection report", "form": [],
            "note_filter": SimpleNamespace(note_filter_form_rows=[], note_filter_choices=[]),
            "result": result, "report_error": "", "filter_summary": "all",
            "calculation_rules": CALCULATION_RULES, "export_query": "export=xlsx",
        }))
        self.assertIn("Transect summary", html)
        self.assertIn("Occurrence detection detail", html)
        for rule in CALCULATION_RULES:
            self.assertIn(rule, html)
