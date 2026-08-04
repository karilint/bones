"""Read-only reconciliation of imported field logs and completed survey data."""
from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _value(item: dict[str, Any], *names: str) -> Any:
    indexed = {_key(k): v for k, v in item.items()}
    for name in names:
        value = indexed.get(_key(name))
        if value not in (None, ""):
            return value
    return None


def _integer(value: Any) -> int | None:
    try:
        return int(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _coordinates(value: Any) -> tuple[float | None, float | None]:
    try:
        latitude, longitude = str(value).split(",", 1)
        return float(latitude), float(longitude)
    except (AttributeError, TypeError, ValueError):
        return None, None


def _datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    parsed = parse_datetime(str(value).replace("Z", "+00:00"))
    if parsed:
        return parsed
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M", "%d-%b-%Y %H:%M:%S"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            pass
    return None


def _children(value: Any) -> Iterable[dict[str, Any]]:
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from _children(child)
    elif isinstance(value, list):
        for child in value:
            yield from _children(child)


def _xml_value(node: ET.Element) -> Any:
    children = list(node)
    if not children:
        return (node.text or "").strip()
    result: dict[str, Any] = dict(node.attrib)
    for child in children:
        value = _xml_value(child)
        if child.tag in result:
            result[child.tag] = result[child.tag] if isinstance(result[child.tag], list) else [result[child.tag]]
            result[child.tag].append(value)
        else:
            result[child.tag] = value
    return result


@dataclass
class ParsedLog:
    log_id: int
    format: str
    transects: list[dict[str, Any]] = field(default_factory=list)
    occurrences: list[dict[str, Any]] = field(default_factory=list)
    instances: list[dict[str, Any]] = field(default_factory=list)
    track_points: list[dict[str, Any]] = field(default_factory=list)
    error: str = ""


def _parse_line_log(log_id: int, contents: str) -> ParsedLog | None:
    """Parse the command-oriented field logger format used by Bones clients."""
    commands = {"STARTTRANSECT", "STARTOCCURRENCE", "STARTWORKFLOW", "QUESTION", "CHECKPOINT"}
    if not any(line.split(" ", 1)[0].upper() in commands for line in contents.splitlines() if line.strip()):
        return None
    result = ParsedLog(log_id, "Bones line log")
    transect = None
    occurrence = None
    workflow = None
    for position, raw_line in enumerate(contents.splitlines(), 1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split()
        command = parts[0].upper()
        if command == "STARTTRANSECT" and len(parts) >= 3:
            lat, long = _coordinates(parts[3] if len(parts) >= 4 else None)
            transect = {"log_id": log_id, "source": f"line {position}", "template": parts[1], "uid": _integer(parts[2]), "name": None, "start_time": _datetime(" ".join(parts[4:6])) if len(parts) >= 6 else None, "end_time": None, "state": "started", "lat_from": lat, "long_from": long, "user": None}
            result.transects.append(transect)
            occurrence = None
            workflow = None
        elif command == "/STARTTRANSECT" and transect:
            transect["user"] = parts[-1] if len(parts) >= 2 else None
        elif command == "CANCELTRANSECT" and transect:
            transect["state"] = "cancelled"
            transect["end_time"] = _datetime(" ".join(parts[1:3]))
            occurrence = None
        elif command == "/ENDTRANSECT" and transect:
            transect["state"] = "completed"
            lat, long = _coordinates(parts[1] if len(parts) >= 2 else None)
            transect["lat_to"], transect["long_to"] = lat, long
            transect["end_time"] = _datetime(" ".join(parts[2:4])) if len(parts) >= 4 else None
            occurrence = None
        elif command == "STARTOCCURRENCE" and transect and len(parts) >= 2:
            lat, long = _coordinates(parts[2] if len(parts) >= 3 else None)
            occurrence = {"log_id": log_id, "source": f"line {position}", "transect_uid": transect["uid"], "id": None, "number": _integer(parts[1]), "start_time": _datetime(" ".join(parts[3:5])) if len(parts) >= 5 else None, "end_time": None, "state": "started", "lat": lat, "long": long, "user": None}
            result.occurrences.append(occurrence)
            workflow = None
        elif command == "/STARTOCCURRENCE" and occurrence:
            occurrence["user"] = parts[-1] if len(parts) >= 2 else None
        elif command == "CANCELOCCURRENCE" and occurrence:
            occurrence["state"] = "cancelled"
            occurrence["end_time"] = _datetime(" ".join(parts[1:3]))
        elif command == "/ENDOCCURRENCE" and occurrence:
            occurrence["state"] = "completed"
            occurrence["end_time"] = _datetime(" ".join(parts[1:3]))
        elif command == "STARTWORKFLOW" and occurrence and len(parts) >= 4:
            workflow = {"log_id": log_id, "source": f"line {position}", "transect_uid": transect["uid"], "occurrence_id": None, "occurrence_number": occurrence["number"], "number": _integer(parts[3]), "workflow_uid": parts[2], "template": parts[1], "parent_occurrence": occurrence, "state": "started", "response_count": 0}
            result.instances.append(workflow)
        elif command == "ENDWORKFLOW" and workflow:
            workflow["state"] = "completed"
        elif command in {"RESPONSE", "SKIPPEDRESPONSE"} and workflow:
            workflow["response_count"] += 1
        elif command in {"CHECKPOINT", "TURNAROUND", "PAUSETRANSECT", "RESUMETRANSECT"} and transect:
            coordinate_token = parts[2] if command == "PAUSETRANSECT" and len(parts) >= 3 else (parts[1] if len(parts) >= 2 else None)
            date_index = 3 if command == "PAUSETRANSECT" else 2
            lat, long = _coordinates(coordinate_token)
            result.track_points.append({"log_id": log_id, "source": f"line {position}", "transect_uid": transect["uid"], "event": command, "time": _datetime(" ".join(parts[date_index:date_index + 2])), "lat": lat, "long": long, "user": transect.get("user")})
    return result


def parse_log(log_id: int, contents: str | None) -> ParsedLog:
    """Parse known JSON/XML log shapes without assuming one serializer version."""
    if not contents or not contents.strip():
        return ParsedLog(log_id, "empty", error="Log contents are empty")
    try:
        root = json.loads(contents)
        fmt = "JSON"
    except (TypeError, json.JSONDecodeError):
        try:
            element = ET.fromstring(contents)
            root = {element.tag: _xml_value(element)}
            fmt = "XML"
        except ET.ParseError as exc:
            line_log = _parse_line_log(log_id, contents)
            if line_log is not None:
                return line_log
            return ParsedLog(log_id, "unknown", error=f"Unsupported or malformed log: {exc}")

    result = ParsedLog(log_id, fmt)
    seen: set[tuple[str, str]] = set()
    for position, item in enumerate(_children(root)):
        uid = _integer(_value(item, "TransectUID", "CompletedTransectUID", "TransectID"))
        occurrence_number = _integer(_value(item, "OccurrenceNumber", "OccurrenceNo"))
        occurrence_id = _integer(_value(item, "OccurrenceID", "CompletedOccurrenceID"))
        instance_number = _integer(_value(item, "InstanceNumber", "InstanceNo"))
        workflow_uid = _value(item, "WorkflowUID", "CompletedWorkflowID", "WorkflowInstanceUID")
        time = _datetime(_value(item, "Time", "RecordedAt", "Timestamp"))
        is_track = any(_value(item, name) is not None for name in ("isStart", "isCheckPoint", "isTurnPoint", "isEnd"))

        if uid is not None and (is_track or (_value(item, "Lat", "Latitude") is not None and time)):
            token = ("track", repr((uid, time, _value(item, "Lat", "Latitude"), _value(item, "Long", "Longitude"))))
            if token not in seen:
                seen.add(token)
                result.track_points.append({"log_id": log_id, "transect_uid": uid, "time": time})

        if uid is not None and any(_value(item, name) is not None for name in ("Name", "TransectName", "StartTime", "EndTime", "TransectTemplateID")):
            token = ("transect", str(uid))
            if token not in seen:
                seen.add(token)
                result.transects.append({
                    "log_id": log_id, "source": f"object {position}", "uid": uid,
                    "name": _value(item, "TransectName", "Name"),
                    "template": _value(item, "TransectTemplateID", "TemplateTransectID", "TemplateName"),
                    "start_time": _datetime(_value(item, "StartTime", "RecordingStartTime")),
                    "end_time": _datetime(_value(item, "EndTime", "RecordingEndTime")),
                    "state": _value(item, "State"),
                })

        if occurrence_number is not None or occurrence_id is not None:
            token = ("occurrence", repr((uid, occurrence_id, occurrence_number)))
            if token not in seen:
                seen.add(token)
                result.occurrences.append({
                    "log_id": log_id, "source": f"object {position}", "transect_uid": uid,
                    "id": occurrence_id, "number": occurrence_number,
                    "start_time": _datetime(_value(item, "RecordingStartTime", "OccurrenceStartTime")),
                })

        if instance_number is not None or workflow_uid is not None:
            token = ("instance", repr((occurrence_id, occurrence_number, instance_number, workflow_uid)))
            if token not in seen:
                seen.add(token)
                result.instances.append({
                    "log_id": log_id, "source": f"object {position}", "transect_uid": uid,
                    "occurrence_id": occurrence_id, "occurrence_number": occurrence_number,
                    "number": instance_number, "workflow_uid": str(workflow_uid or ""),
                    "template": _value(item, "TemplateWorkflowID", "WorkflowTemplateID", "WorkflowName"),
                })
    if not (result.transects or result.occurrences or result.instances or result.track_points):
        result.error = "Recognized serialization but found no reconciliation entities"
    return result


def gps_status(year: int | None, db_points: int, log_points: int, required_from: int | None) -> str:
    if db_points:
        return "GPS_PRESENT"
    if log_points:
        return "GPS_MISSING"
    if required_from is None:
        return "GPS_EXPECTATION_UNKNOWN"
    if year is not None and year < required_from:
        return "GPS_NOT_EXPECTED_EARLY_MANUAL"
    return "GPS_MISSING"


SHEETS = {
    "Summary": ("Metric", "Value"),
    "Critical findings": ("Severity", "Entity", "Status", "Log ID", "Identifier", "Reason", "Evidence"),
    "Transects": ("Log ID", "Logged UID", "Database UID", "Name", "Year", "Status", "Confidence", "Occurrences log", "Occurrences DB", "GPS status", "Deletion evidence", "Reason", "Field label"),
    "Occurrences": ("Log ID", "Transect UID", "Logged occurrence ID", "Occurrence number", "Database ID", "Status", "Confidence", "Instances log", "Instances DB", "Deletion evidence", "Reason", "Field label"),
    "Instances": ("Log ID", "Transect UID", "Occurrence ID", "Occurrence number", "Instance number", "Workflow UID", "Database workflow UIDs", "Status", "Confidence", "Responses DB", "Deletion evidence", "Reason", "Field label"),
    "GPS": ("Transect UID", "Name", "Year", "Database points", "Log points", "Status", "Reason", "Field label"),
    "Recovery candidates": ("Candidate type", "Field label", "Recovery status", "Confidence", "Log ID", "Source", "Parent", "Import order", "Existing match", "Deletion/history evidence", "Blocking reason", "Event time", "Latitude", "Longitude", "Source identifier"),
    "Database only": ("Entity", "Identifier", "Parent", "Reason"),
    "Deleted evidence": ("Entity", "Identifier", "Parent", "Deleted at", "Reason", "Source"),
    "Log parse issues": ("Log ID", "Upload date", "Uploaded by", "Format", "Error"),
    "Methodology": ("Topic", "Description"),
}


def workbook_bytes(rows: dict[str, list[list[Any]]], included_sheets: Iterable[str] | None = None) -> bytes:
    """Build a reconciliation workbook for web downloads or file output."""
    selected = set(included_sheets or SHEETS)
    selected.add("Summary")
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, headers in SHEETS.items():
        if title not in selected:
            continue
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in rows.get(title, []):
            sheet.append([_excel(value) for value in row])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, len(headers) + 1):
            values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
            sheet.column_dimensions[get_column_letter(column)].width = min(60, max(12, max(map(len, values)) + 2))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_workbook(path: str | Path, rows: dict[str, list[list[Any]]], included_sheets: Iterable[str] | None = None) -> Path:
    target = Path(path).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(workbook_bytes(rows, included_sheets))
    return target


def _excel(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo:
        return value.replace(tzinfo=None)
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, default=str, ensure_ascii=False)
    return value


def summary_rows(rows: dict[str, list[list[Any]]]) -> list[list[Any]]:
    status_columns = {"Transects": 5, "Occurrences": 5, "Instances": 7}
    counts = Counter(
        row[status_columns[name]]
        for name in status_columns
        for row in rows.get(name, [])
    )
    return [["Generated", datetime.now().isoformat(timespec="seconds")], ["Critical findings", len(rows.get("Critical findings", []))], ["Log parse issues", len(rows.get("Log parse issues", []))], *[[f"Status: {key}", value] for key, value in sorted(counts.items())]]
