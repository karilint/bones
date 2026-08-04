from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .models import CompletedOccurrenceInfo, DataTypeOption


HEADERS = (
    "completed_transect_name", "occurrence_number", "pre_or_post",
    "question_text", "new_response", "update_comment",
)
MAX_ROWS = 5000
COLUMN_GUIDE = (
    ("completed_transect_name", "Yes", "Visible completed transect name; not the template name, UID, or database key.", "1", "Exact name match, ignoring case and outer spaces."),
    ("occurrence_number", "Yes", "Occurrence number shown within the transect; not the occurrence database key.", "5", "Positive whole number."),
    ("pre_or_post", "Yes", "Whether the occurrence information is recorded before or after workflows.", "Post", "Only Pre or Post."),
    ("question_text", "Yes", "Complete question text displayed to users; not a question ID.", "Taxon Guess?", "Exact text match, ignoring case and outer spaces."),
    ("new_response", "Yes", "Visible answer text to store; use displayed option text, not its response code.", "waterbuck", "Must match a configured option when options exist."),
    ("update_comment", "Yes", "Short reason for the change, stored in history when the answer changes.", "Taxon identification corrected.", "Required; maximum 100 characters."),
)


def template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Updates"
    sheet.append(HEADERS)
    sheet.append(("1", 5, "Post", "Taxon Guess?", "waterbuck", "Taxon identification corrected."))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F2"
    for index, width in enumerate((26, 20, 14, 34, 28, 48), 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="417690")
    for cell, guide in zip(sheet[1], COLUMN_GUIDE):
        cell.comment = Comment(guide[2], "Bones")
    phase = DataValidation(type="list", formula1='"Pre,Post"', allow_blank=False)
    sheet.add_data_validation(phase)
    phase.add("C2:C5001")
    comment = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="100", allow_blank=False)
    comment.error = "Update comments are required and may contain at most 100 characters."
    comment.errorTitle = "Invalid update comment"
    comment.showErrorMessage = True
    sheet.add_data_validation(comment)
    comment.add("F2:F5001")
    instructions = workbook.create_sheet("Instructions")
    instructions.append(("Occurrence answer bulk update",))
    instructions.append(("All input identifiers are human-readable values. Use the completed transect name, not its template name, database UID, or key.",))
    instructions.append(("Enter one update per row on the Updates sheet and delete the example row first.",))
    instructions.append(("Uploading creates a preview. Review the resolved transect and current answer before confirming.",))
    instructions.append(("Changed answers create history with the update comment. Unchanged and duplicate rows create no history and their comments are not stored.",))
    instructions.append(("Ambiguous or missing matches and comments over 100 characters are errors; comments are never truncated.",))
    instructions.append(())
    instructions.append(("Column", "Required", "Human-readable value expected", "Example", "Validation"))
    for guide in COLUMN_GUIDE:
        instructions.append(guide)
    instructions.freeze_panes = "A9"
    instructions.auto_filter.ref = "A8:E14"
    for cell in instructions[8]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="417690")
    for column, width in zip(("A", "B", "C", "D", "E"), (28, 12, 70, 38, 55)):
        instructions.column_dimensions[column].width = width
    instructions["A1"].font = Font(bold=True, size=14)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _error(row_number, values, message):
    return {"row_number": row_number, **values, "status": "error", "message": message}


def validate_workbook(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise ValueError("The uploaded file is not a readable Excel workbook.") from exc
    if "Updates" not in workbook.sheetnames:
        raise ValueError("The workbook must contain a sheet named 'Updates'.")
    rows = workbook["Updates"].iter_rows(values_only=True)
    try:
        headers = tuple(_text(value) for value in next(rows))
    except StopIteration as exc:
        raise ValueError("The Updates sheet is empty.") from exc
    if headers != HEADERS:
        raise ValueError("The Updates headings must exactly match the downloadable template.")

    results = []
    for row_number, row in enumerate(rows, 2):
        if row_number > MAX_ROWS + 1:
            raise ValueError(f"The workbook may contain at most {MAX_ROWS} update rows.")
        if not any(value is not None and _text(value) for value in row):
            continue
        values = dict(zip(HEADERS, (_text(value) for value in row)))
        if any(str(value).startswith("=") for value in row if value is not None):
            results.append(_error(row_number, values, "Formulas are not allowed in import rows.")); continue
        missing = [name for name in HEADERS if not values[name]]
        if missing:
            results.append(_error(row_number, values, f"Required value missing: {', '.join(missing)}.")); continue
        try:
            number = int(values["occurrence_number"])
            if number < 1 or str(number) != values["occurrence_number"].removesuffix(".0"):
                raise ValueError
        except ValueError:
            results.append(_error(row_number, values, "Occurrence number must be a positive whole number.")); continue
        phase = values["pre_or_post"].casefold()
        if phase not in {"pre", "post"}:
            results.append(_error(row_number, values, "Pre/Post must be either Pre or Post.")); continue
        values["pre_or_post"] = phase.title()
        if len(values["update_comment"]) > 100:
            results.append(_error(row_number, values, "Update comment exceeds 100 characters.")); continue
        matches = list(CompletedOccurrenceInfo.objects.select_related(
            "occurrence__transect__transect_template"
        ).filter(
            occurrence__transect__name__iexact=values["completed_transect_name"],
            occurrence__occurrence_number=number,
            pre_or_post__iexact=values["pre_or_post"],
            question_text__iexact=values["question_text"],
        )[:2])
        if not matches:
            results.append(_error(row_number, values, "No matching occurrence-info answer was found.")); continue
        if len(matches) > 1:
            results.append(_error(row_number, values, "More than one occurrence-info answer matched these values.")); continue
        target = matches[0]
        new_code = ""
        new_response = values["new_response"]
        if target.response_data_type:
            options = list(DataTypeOption.objects.filter(
                data_type_id=target.response_data_type, text__iexact=new_response,
            )[:2])
            if len(options) > 1:
                results.append(_error(row_number, values, "More than one configured answer option matched the new response.")); continue
            if len(options) == 1:
                new_code = str(options[0].code or "")
                new_response = options[0].text or ""
            elif DataTypeOption.objects.filter(data_type_id=target.response_data_type).exists():
                results.append(_error(row_number, values, "The new response is not a configured option for this question.")); continue
        unchanged = (
            (target.response_code or "") == new_code
            and (target.response or "").strip().casefold() == new_response.strip().casefold()
        )
        results.append({
            "row_number": row_number, **values,
            "status": "unchanged" if unchanged else "ready",
            "message": (
                "Answer already has the requested value. No update was made and no history "
                "entry was created; the import comment was not stored."
                if unchanged else "Ready to update."
            ),
            "target_id": target.pk,
            "transect_uid": target.occurrence.transect_id,
            "transect_name": target.occurrence.transect.name,
            "occurrence_id": target.occurrence_id,
            "current_response_code": target.response_code or "",
            "current_response": target.response or "",
            "new_response_code": new_code,
            "canonical_new_response": new_response,
        })

    seen = {}
    for item in results:
        if item["status"] not in {"ready", "unchanged"}:
            continue
        target_id = str(item["target_id"])
        previous = seen.get(target_id)
        if previous is None:
            seen[target_id] = item
            continue
        same = (
            previous["new_response_code"], previous["canonical_new_response"], previous["update_comment"]
        ) == (
            item["new_response_code"], item["canonical_new_response"], item["update_comment"]
        )
        if same:
            item["status"] = "duplicate"
            item["message"] = (
                f"Duplicate of row {previous['row_number']}. No additional update or history "
                "entry was created; the import comment was not stored."
            )
        else:
            previous["status"] = "error"
            previous["message"] = f"Conflicts with another update for the same answer (row {item['row_number']})."
            item["status"] = "error"
            item["message"] = f"Conflicts with another update for the same answer (row {previous['row_number']})."
    return results
